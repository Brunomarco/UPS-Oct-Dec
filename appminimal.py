import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import heapq
from io import BytesIO

# PDF generation imports
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

def generate_routes_excel(routes, origin, destination, selected_date, route_type="fastest"):
    """
    Generate a professionally formatted Excel file with route information.
    
    Args:
        routes: List of route dictionaries
        origin: Origin airport code
        destination: Destination airport code
        selected_date: Selected departure date
        route_type: "fastest" or "fewest_stops"
    
    Returns:
        bytes containing the Excel file
    """
    from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import FormulaRule
    
    buffer = BytesIO()
    
    # Define UPS brand colors
    ups_brown = "351C15"
    ups_gold = "FFB500"
    light_brown = "F5E6D3"
    light_gray = "F2F2F2"
    white = "FFFFFF"
    
    # Define styles
    header_font = Font(bold=True, color=white, size=11)
    header_fill = PatternFill(start_color=ups_brown, end_color=ups_brown, fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    data_font = Font(size=10)
    data_alignment = Alignment(horizontal="left", vertical="center")
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    alt_row_fill = PatternFill(start_color=light_gray, end_color=light_gray, fill_type="solid")
    route_separator_fill = PatternFill(start_color=light_brown, end_color=light_brown, fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Prepare data for Excel - extract all flight segments
    all_segments = []
    
    for route_idx, route in enumerate(routes, 1):
        route_str = " → ".join(route['path'])
        total_hours = route['total_duration'] // 60
        total_mins = route['total_duration'] % 60
        
        for seg_idx, leg in enumerate(route['route_info'], 1):
            # Calculate arrival date (handle overnight flights)
            leg_departure_date = leg['date']
            leg_arrival_date = leg['date']
            
            try:
                dep_str = str(leg['departure'])
                arr_str = str(leg['arrival'])
                if ':' in dep_str and ':' in arr_str:
                    dep_parts = dep_str.split(':')
                    arr_parts = arr_str.split(':')
                    dep_minutes = int(dep_parts[0]) * 60 + int(dep_parts[1])
                    arr_minutes = int(arr_parts[0]) * 60 + int(arr_parts[1])
                    if arr_minutes < dep_minutes:
                        leg_arrival_date = leg['date'] + timedelta(days=1)
            except:
                pass
            
            # Connection time to next segment
            if seg_idx < len(route['route_info']):
                next_wait = route['route_info'][seg_idx]['wait_time']
                conn_hours = next_wait // 60
                conn_mins = next_wait % 60
                connection_str = f"{conn_hours}h {conn_mins}m"
            else:
                connection_str = "—"
            
            segment_data = {
                'Route #': route_idx,
                'Route': route_str,
                'Stops': route['stops'],
                'Segment': f"{seg_idx} of {len(route['route_info'])}",
                'Carrier': leg['carrier'],
                'Flight': leg['flight'],
                'Origin': leg['from'],
                'Destination': leg['to'],
                'Dep. Date': leg_departure_date.strftime('%Y-%m-%d'),
                'Dep. Day': leg_departure_date.strftime('%a'),
                'Dep. Time': leg['departure'],
                'Arr. Date': leg_arrival_date.strftime('%Y-%m-%d'),
                'Arr. Day': leg_arrival_date.strftime('%a'),
                'Arr. Time': leg['arrival'],
                'Flight Duration': leg['duration_str'],
                'Connection': connection_str,
                'Total Journey': f"{total_hours}h {total_mins}m",
                'Final Arrival': route['arrival_datetime'].strftime('%Y-%m-%d %H:%M')
            }
            all_segments.append(segment_data)
    
    # Create DataFrame
    df = pd.DataFrame(all_segments)
    
    # Write to Excel with formatting
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        # Write Route Details sheet
        df.to_excel(writer, sheet_name='Route Details', index=False, startrow=2)
        
        workbook = writer.book
        worksheet = writer.sheets['Route Details']
        
        # Add title row
        route_type_display = 'Fewest Stops & Shortest Journey' if route_type == 'fewest_stops' else 'Fastest Arriving'
        title = f"UPS Healthcare Logistics - Flight Routing: {origin} → {destination}"
        subtitle = f"Departure Date: {selected_date} | Route Type: {route_type_display} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        
        worksheet.merge_cells('A1:R1')
        worksheet.merge_cells('A2:R2')
        
        title_cell = worksheet['A1']
        title_cell.value = title
        title_cell.font = Font(bold=True, size=14, color=ups_brown)
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        
        subtitle_cell = worksheet['A2']
        subtitle_cell.value = subtitle
        subtitle_cell.font = Font(size=10, color="666666")
        subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Style header row (row 3)
        for col_idx, col in enumerate(df.columns, 1):
            cell = worksheet.cell(row=3, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Style data rows
        prev_route = None
        for row_idx in range(4, len(df) + 4):
            current_route = worksheet.cell(row=row_idx, column=1).value
            
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                
                # Center alignment for specific columns
                if col_idx in [1, 3, 4, 5, 6, 10, 13]:  # Route #, Stops, Segment, Carrier, Flight, Days
                    cell.alignment = center_alignment
                else:
                    cell.alignment = data_alignment
                
                # Alternating row colors within each route group
                if current_route != prev_route and prev_route is not None:
                    # First row of new route - light brown
                    cell.fill = route_separator_fill
                elif (row_idx - 4) % 2 == 1:
                    cell.fill = alt_row_fill
            
            prev_route = current_route
        
        # Set column widths
        column_widths = {
            'A': 8,   # Route #
            'B': 25,  # Route
            'C': 7,   # Stops
            'D': 10,  # Segment
            'E': 8,   # Carrier
            'F': 10,  # Flight
            'G': 8,   # Origin
            'H': 10,  # Destination
            'I': 12,  # Dep. Date
            'J': 6,   # Dep. Day
            'K': 10,  # Dep. Time
            'L': 12,  # Arr. Date
            'M': 6,   # Arr. Day
            'N': 10,  # Arr. Time
            'O': 14,  # Flight Duration
            'P': 12,  # Connection
            'Q': 14,  # Total Journey
            'R': 18,  # Final Arrival
        }
        
        for col_letter, width in column_widths.items():
            worksheet.column_dimensions[col_letter].width = width
        
        # Freeze panes (freeze header row)
        worksheet.freeze_panes = 'A4'
        
        # Add auto-filter
        worksheet.auto_filter.ref = f"A3:R{len(df) + 3}"
        
        # Set row height for header
        worksheet.row_dimensions[1].height = 25
        worksheet.row_dimensions[2].height = 18
        worksheet.row_dimensions[3].height = 35
        
        # ============================================================
        # Create Summary sheet
        # ============================================================
        summary_ws = workbook.create_sheet('Summary')
        
        # Title
        summary_ws.merge_cells('A1:C1')
        summary_ws['A1'] = "Route Summary"
        summary_ws['A1'].font = Font(bold=True, size=14, color=ups_brown)
        summary_ws['A1'].alignment = Alignment(horizontal="left")
        
        # Summary data
        summary_info = [
            ('Origin Airport:', origin),
            ('Destination Airport:', destination),
            ('Departure Date:', str(selected_date)),
            ('Route Type:', route_type_display),
            ('Total Routes Found:', len(routes)),
            ('Report Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ('', ''),
            ('Route Overview:', ''),
        ]
        
        for row_idx, (label, value) in enumerate(summary_info, 3):
            summary_ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True, size=10)
            summary_ws.cell(row=row_idx, column=2, value=value).font = Font(size=10)
        
        # Add route summary table
        route_summary_start = len(summary_info) + 4
        
        # Headers for route summary
        route_headers = ['Route #', 'Route', 'Stops', 'Total Journey', 'Final Arrival']
        for col_idx, header in enumerate(route_headers, 1):
            cell = summary_ws.cell(row=route_summary_start, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Route data
        for route_idx, route in enumerate(routes, 1):
            row = route_summary_start + route_idx
            route_str = " → ".join(route['path'])
            total_hours = route['total_duration'] // 60
            total_mins = route['total_duration'] % 60
            
            data = [
                route_idx,
                route_str,
                route['stops'],
                f"{total_hours}h {total_mins}m",
                route['arrival_datetime'].strftime('%Y-%m-%d %H:%M')
            ]
            
            for col_idx, value in enumerate(data, 1):
                cell = summary_ws.cell(row=row, column=col_idx, value=value)
                cell.font = data_font
                cell.border = thin_border
                if col_idx in [1, 3]:
                    cell.alignment = center_alignment
                else:
                    cell.alignment = data_alignment
                
                if route_idx % 2 == 0:
                    cell.fill = alt_row_fill
        
        # Set column widths for summary
        summary_ws.column_dimensions['A'].width = 10
        summary_ws.column_dimensions['B'].width = 30
        summary_ws.column_dimensions['C'].width = 8
        summary_ws.column_dimensions['D'].width = 15
        summary_ws.column_dimensions['E'].width = 20
    
    buffer.seek(0)
    return buffer.getvalue()

def generate_routes_pdf(routes, origin, destination, selected_date, route_type="fastest"):
    """
    Generate a professional PDF document with route information.
    
    Args:
        routes: List of route dictionaries
        origin: Origin airport code
        destination: Destination airport code
        selected_date: Selected departure date
        route_type: "fastest" or "fewest_stops"
    
    Returns:
        bytes containing the PDF
    """
    buffer = BytesIO()
    
    # Create document
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch
    )
    
    # Define styles
    styles = getSampleStyleSheet()
    
    # UPS Brand colors
    ups_brown = colors.HexColor('#351C15')
    ups_gold = colors.HexColor('#FFB500')
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=ups_brown,
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=ups_brown,
        spaceAfter=10,
        alignment=TA_CENTER,
        fontName='Helvetica'
    )
    
    header_style = ParagraphStyle(
        'CustomHeader',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=ups_brown,
        spaceBefore=15,
        spaceAfter=10,
        fontName='Helvetica-Bold'
    )
    
    subheader_style = ParagraphStyle(
        'CustomSubHeader',
        parent=styles['Heading3'],
        fontSize=12,
        textColor=ups_brown,
        spaceBefore=10,
        spaceAfter=5,
        fontName='Helvetica-Bold'
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.black,
        spaceAfter=5,
        fontName='Helvetica'
    )
    
    small_style = ParagraphStyle(
        'CustomSmall',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.grey,
        spaceAfter=3,
        fontName='Helvetica'
    )
    
    # Build story (content)
    story = []
    
    # Header section
    story.append(Paragraph("UPS Healthcare Logistics", title_style))
    story.append(Paragraph(f"Routing {origin} to {destination} on {selected_date}", subtitle_style))
    story.append(Spacer(1, 10))
    
    # Route type header
    if route_type == "fastest":
        route_type_text = "Fastest Arriving Routes"
        route_type_desc = "Routes sorted by earliest arrival at destination"
    else:
        route_type_text = "Routes with Fewest Stops"
        route_type_desc = "Routes with minimum number of connections"
    
    story.append(Paragraph(f"<b>{route_type_text}</b>", header_style))
    story.append(Paragraph(route_type_desc, small_style))
    story.append(Spacer(1, 5))
    
    # Summary info table
    summary_data = [
        ['Origin:', origin, 'Destination:', destination],
        ['Departure Date:', str(selected_date), 'Generated:', datetime.now().strftime('%Y-%m-%d %H:%M')],
        ['Total Routes:', str(len(routes)), '', '']
    ]
    
    summary_table = Table(summary_data, colWidths=[1.2*inch, 1.8*inch, 1.2*inch, 1.8*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#FFF8E8')),
        ('TEXTCOLOR', (0, 0), (-1, -1), ups_brown),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('BOX', (0, 0), (-1, -1), 1, ups_brown),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 20))
    
    # Generate page for each route
    for idx, route in enumerate(routes, 1):
        if idx > 1:
            story.append(PageBreak())
        
        route_str = " → ".join(route['path'])
        total_duration = route['total_duration']
        total_hours = total_duration // 60
        total_mins = total_duration % 60
        
        total_wait = sum([leg['wait_time'] for leg in route['route_info']])
        wait_hours = total_wait // 60
        wait_mins = total_wait % 60
        
        first_leg = route['route_info'][0]
        dep_time_str = first_leg['departure']
        
        # Route header
        story.append(Paragraph(f"Route {idx}: {route_str}", header_style))
        
        # Route summary box
        if route_type == "fastest":
            box_color = colors.HexColor('#E8F4F8')
            label = "Fastest Arriving"
        else:
            box_color = colors.HexColor('#E8F8E8')
            label = "Fewest Stops"
        
        route_summary_data = [
            [f'Route Summary - {label}', ''],
            ['Route:', route_str],
            ['Number of Stops:', f"{route['stops']} stop(s)"],
            ['Departure:', f"{route['start_date'].strftime('%Y-%m-%d')} at {dep_time_str} ({route['start_date'].strftime('%A')})"],
            ['Arrival:', f"{route['arrival_datetime'].strftime('%Y-%m-%d')} at {route['arrival_datetime'].strftime('%H:%M')} ({route['arrival_datetime'].strftime('%A')})"],
            ['Total Journey Time:', f"{total_hours}h {total_mins}m"],
            ['Total Waiting Time:', f"{wait_hours}h {wait_mins}m"],
        ]
        
        route_table = Table(route_summary_data, colWidths=[2*inch, 4.5*inch])
        route_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), ups_brown),
            ('TEXTCOLOR', (0, 0), (-1, 0), ups_gold),
            ('BACKGROUND', (0, 1), (-1, -1), box_color),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('SPAN', (0, 0), (1, 0)),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 10),
            ('BOX', (0, 0), (-1, -1), 1, ups_brown),
            ('LINEBELOW', (0, 0), (-1, 0), 1, ups_brown),
        ]))
        story.append(route_table)
        story.append(Spacer(1, 15))
        
        # Flight segments header
        story.append(Paragraph("Flight Segments", subheader_style))
        
        # Flight segments table
        segments_header = ['Seg', 'From', 'To', 'Date', 'Departure', 'Arrival', 'Duration', 'Carrier', 'Flight', 'Connection']
        segments_data = [segments_header]
        
        for j, leg in enumerate(route['route_info'], 1):
            leg_departure_date = leg['date']
            leg_arrival_date = leg['date']
            
            # Check for overnight flight
            dep_str = leg['departure']
            arr_str = leg['arrival']
            
            try:
                dep_parts = dep_str.split(':')
                arr_parts = arr_str.split(':')
                dep_minutes = int(dep_parts[0]) * 60 + int(dep_parts[1])
                arr_minutes = int(arr_parts[0]) * 60 + int(arr_parts[1])
                if arr_minutes < dep_minutes:
                    leg_arrival_date = leg['date'] + timedelta(days=1)
            except:
                pass
            
            # Connection time
            if j < len(route['route_info']):
                next_wait = route['route_info'][j]['wait_time']
                conn_hours = next_wait // 60
                conn_mins = next_wait % 60
                connection_str = f"{conn_hours}h {conn_mins}m"
            else:
                connection_str = "Final"
            
            # Duration
            dur_hours = leg['duration'] // 60
            dur_mins = leg['duration'] % 60
            duration_str = f"{dur_hours}h {dur_mins}m"
            
            row = [
                str(j),
                leg['from'],
                leg['to'],
                leg_departure_date.strftime('%Y-%m-%d'),
                f"{leg['departure']}",
                f"{leg['arrival']}",
                duration_str,
                leg['carrier'],
                leg['flight'],
                connection_str
            ]
            segments_data.append(row)
        
        # Create segments table with proper column widths
        col_widths = [0.35*inch, 0.5*inch, 0.5*inch, 0.85*inch, 0.65*inch, 0.65*inch, 0.65*inch, 0.55*inch, 0.65*inch, 0.75*inch]
        segments_table = Table(segments_data, colWidths=col_widths)
        
        segments_table.setStyle(TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), ups_brown),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            # Data rows
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#FAFAFA')),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            # Alignment
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            # Padding
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            # Borders
            ('BOX', (0, 0), (-1, -1), 1, ups_brown),
            ('LINEBELOW', (0, 0), (-1, 0), 1, ups_brown),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ]))
        
        story.append(segments_table)
        story.append(Spacer(1, 15))
        
        # Warning for 5+ stops
        if route['stops'] >= 5:
            warning_data = [['⚠️ Complex Routing - This route requires 5 or more stops. Please contact UPS Healthcare Logistics for assistance.']]
            warning_table = Table(warning_data, colWidths=[6.5*inch])
            warning_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#FFF3CD')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#856404')),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('TOPPADDING', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#FF6B00')),
            ]))
            story.append(warning_table)
    
    # Footer on last page
    story.append(Spacer(1, 30))
    footer_text = f"Generated by UPS Healthcare Logistics Flight Routing System | {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    story.append(Paragraph(footer_text, small_style))
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    
    return buffer.getvalue()

# Page configuration with UPS branding
st.set_page_config(
    page_title="UPS Flight Routing System",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Initialize session state for tab persistence
if 'selected_tab' not in st.session_state:
    st.session_state.selected_tab = "Tracked Routes"

# Initialize session state for caching route results (prevents reload on download)
if 'cached_routes' not in st.session_state:
    st.session_state.cached_routes = None
if 'cached_origin' not in st.session_state:
    st.session_state.cached_origin = None
if 'cached_destination' not in st.session_state:
    st.session_state.cached_destination = None
if 'cached_date' not in st.session_state:
    st.session_state.cached_date = None

# ============================================================================
# ENTERPRISE CSS - UPS Healthcare Logistics Dashboard
# Professional Grade Design for Global Deployment
# ============================================================================
st.markdown("""
    <style>
    /* ============================================
       TYPOGRAPHY - IBM Plex Sans (Enterprise Standard)
       ============================================ */
    @import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Sans:wght@400;500;600;700&display=swap');
    
    /* ============================================
       DESIGN TOKENS
       ============================================ */
    :root {
        /* UPS Brand Colors */
        --ups-brown-900: #1a0e0a;
        --ups-brown-800: #2d1810;
        --ups-brown-700: #351C15;
        --ups-brown-600: #4a2a1f;
        --ups-brown-500: #5c372a;
        --ups-gold-500: #FFB500;
        --ups-gold-400: #ffc233;
        --ups-gold-600: #d99a00;
        
        /* Neutral Palette */
        --neutral-50: #fafbfc;
        --neutral-100: #f4f5f7;
        --neutral-200: #e8eaed;
        --neutral-300: #dfe1e6;
        --neutral-400: #c1c7d0;
        --neutral-500: #97a0af;
        --neutral-600: #6b778c;
        --neutral-700: #505f79;
        --neutral-800: #344563;
        --neutral-900: #172b4d;
        
        /* Semantic Colors */
        --success-50: #e3fcef;
        --success-500: #00875a;
        --success-600: #006644;
        --warning-50: #fffae6;
        --warning-500: #ff991f;
        --error-50: #ffebe6;
        --error-500: #de350b;
        --info-50: #e6fcff;
        --info-500: #0065ff;
        
        /* Spacing Scale */
        --space-1: 4px;
        --space-2: 8px;
        --space-3: 12px;
        --space-4: 16px;
        --space-5: 20px;
        --space-6: 24px;
        --space-8: 32px;
        --space-10: 40px;
        
        /* Border Radius */
        --radius-sm: 4px;
        --radius-md: 6px;
        --radius-lg: 8px;
        --radius-xl: 12px;
        
        /* Shadows */
        --shadow-sm: 0 1px 2px rgba(23, 43, 77, 0.04);
        --shadow-md: 0 1px 3px rgba(23, 43, 77, 0.1), 0 1px 2px rgba(23, 43, 77, 0.06);
        --shadow-lg: 0 4px 6px rgba(23, 43, 77, 0.1), 0 2px 4px rgba(23, 43, 77, 0.06);
        --shadow-xl: 0 10px 20px rgba(23, 43, 77, 0.1), 0 3px 6px rgba(23, 43, 77, 0.05);
    }
    
    /* ============================================
       GLOBAL RESET & BASE STYLES
       ============================================ */
    .main .block-container {
        padding: var(--space-6) var(--space-8);
        max-width: 1440px;
        background: var(--neutral-50);
    }
    
    html, body, [class*="css"] {
        font-family: 'IBM Plex Sans', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
        font-size: 14px;
        line-height: 1.5;
        color: var(--neutral-900);
        -webkit-font-smoothing: antialiased;
    }
    
    h1, h2, h3, h4, h5, h6 {
        font-family: 'IBM Plex Sans', -apple-system, BlinkMacSystemFont, sans-serif;
        font-weight: 600;
        color: var(--neutral-900);
        letter-spacing: -0.01em;
        margin: 0;
    }
    
    h1 { font-size: 24px; line-height: 1.25; }
    h2 { font-size: 20px; line-height: 1.3; }
    h3 { font-size: 16px; line-height: 1.4; }
    h4 { font-size: 14px; line-height: 1.4; }
    
    p { margin: 0 0 var(--space-3) 0; }
    
    /* ============================================
       SIDEBAR - PROFESSIONAL DARK PANEL
       ============================================ */
    section[data-testid="stSidebar"] {
        background: var(--ups-brown-900);
        border-right: 1px solid var(--ups-brown-800);
    }
    
    section[data-testid="stSidebar"] > div:first-child {
        padding: var(--space-5);
    }
    
    section[data-testid="stSidebar"] h1,
    section[data-testid="stSidebar"] h2,
    section[data-testid="stSidebar"] h3,
    section[data-testid="stSidebar"] h4,
    section[data-testid="stSidebar"] label,
    section[data-testid="stSidebar"] .stMarkdown,
    section[data-testid="stSidebar"] p,
    section[data-testid="stSidebar"] span {
        color: rgba(255, 255, 255, 0.92) !important;
    }
    
    /* Sidebar File Uploader */
    section[data-testid="stSidebar"] [data-testid="stFileUploader"] {
        background: rgba(255, 255, 255, 0.04);
        border: 1px dashed rgba(255, 181, 0, 0.35);
        border-radius: var(--radius-lg);
        padding: var(--space-4);
        transition: all 0.2s ease;
    }
    
    section[data-testid="stSidebar"] [data-testid="stFileUploader"]:hover {
        background: rgba(255, 181, 0, 0.08);
        border-color: var(--ups-gold-500);
    }
    
    section[data-testid="stSidebar"] [data-testid="stFileUploader"] button {
        background: var(--ups-gold-500) !important;
        color: var(--ups-brown-700) !important;
        font-weight: 600 !important;
        border: none !important;
        border-radius: var(--radius-md) !important;
        padding: var(--space-2) var(--space-4) !important;
    }
    
    section[data-testid="stSidebar"] [data-testid="stFileUploader"] small {
        color: rgba(255, 255, 255, 0.5) !important;
    }
    
    section[data-testid="stSidebar"] .stSuccess {
        background: rgba(0, 135, 90, 0.15) !important;
        border: 1px solid rgba(0, 135, 90, 0.3) !important;
        border-radius: var(--radius-md) !important;
    }
    
    /* ============================================
       MAIN HEADER - ENTERPRISE BRANDING
       ============================================ */
    .app-header {
        background: var(--ups-brown-700);
        border-radius: var(--radius-lg);
        padding: 0;
        margin-bottom: var(--space-6);
        overflow: hidden;
        box-shadow: var(--shadow-lg);
    }
    
    .app-header-inner {
        display: flex;
        align-items: center;
        padding: var(--space-4) var(--space-6);
        border-top: 3px solid var(--ups-gold-500);
    }
    
    .app-header-logo {
        background: white;
        padding: var(--space-2) var(--space-3);
        border-radius: var(--radius-md);
        margin-right: var(--space-6);
        box-shadow: var(--shadow-sm);
    }
    
    .app-header-logo img {
        height: 40px;
        width: auto;
        display: block;
    }
    
    .app-header-content {
        flex: 1;
    }
    
    .app-header-title {
        color: var(--ups-gold-500) !important;
        font-size: 20px !important;
        font-weight: 600 !important;
        margin: 0 0 2px 0 !important;
        letter-spacing: -0.02em;
    }
    
    .app-header-subtitle {
        color: rgba(255, 255, 255, 0.7);
        font-size: 13px;
        font-weight: 400;
        margin: 0;
    }
    
    .app-header-meta {
        display: flex;
        align-items: center;
        gap: var(--space-4);
    }
    
    .status-indicator {
        display: flex;
        align-items: center;
        gap: var(--space-2);
        background: rgba(0, 135, 90, 0.15);
        border: 1px solid rgba(0, 135, 90, 0.25);
        padding: var(--space-1) var(--space-3);
        border-radius: 100px;
        font-size: 12px;
        font-weight: 500;
        color: #57d9a3;
    }
    
    .status-indicator::before {
        content: '';
        width: 6px;
        height: 6px;
        background: #36b37e;
        border-radius: 50%;
    }
    
    .marken-badge {
        font-size: 11px;
        font-weight: 600;
        color: rgba(255, 255, 255, 0.5);
        text-transform: uppercase;
        letter-spacing: 0.05em;
    }
    
    /* ============================================
       METRIC CARDS - DASHBOARD KPIs
       ============================================ */
    .kpi-card {
        background: white;
        border: 1px solid var(--neutral-200);
        border-radius: var(--radius-lg);
        padding: var(--space-5);
        box-shadow: var(--shadow-sm);
        transition: all 0.2s ease;
        position: relative;
    }
    
    .kpi-card::after {
        content: '';
        position: absolute;
        left: 0;
        top: 0;
        bottom: 0;
        width: 3px;
        background: var(--ups-gold-500);
        border-radius: var(--radius-lg) 0 0 var(--radius-lg);
    }
    
    .kpi-card:hover {
        border-color: var(--neutral-300);
        box-shadow: var(--shadow-md);
    }
    
    .kpi-value {
        font-size: 28px;
        font-weight: 700;
        color: var(--ups-brown-700);
        line-height: 1;
        margin-bottom: var(--space-1);
    }
    
    .kpi-label {
        font-size: 12px;
        font-weight: 500;
        color: var(--neutral-600);
        text-transform: uppercase;
        letter-spacing: 0.04em;
    }
    
    /* ============================================
       NAVIGATION TABS - SEGMENTED CONTROL
       ============================================ */
    .nav-tabs-container {
        background: var(--neutral-100);
        border: 1px solid var(--neutral-200);
        border-radius: var(--radius-lg);
        padding: var(--space-1);
        margin-bottom: var(--space-5);
    }
    
    div[data-testid="stHorizontalBlock"] .stRadio > div {
        background: transparent;
        padding: 0;
        gap: var(--space-1);
        border: none;
    }
    
    div[data-testid="stHorizontalBlock"] .stRadio > div > label {
        background: transparent;
        padding: var(--space-3) var(--space-5);
        border-radius: var(--radius-md);
        font-weight: 500;
        font-size: 13px;
        color: var(--neutral-700);
        transition: all 0.15s ease;
        border: 1px solid transparent;
    }
    
    div[data-testid="stHorizontalBlock"] .stRadio > div > label:hover {
        background: white;
        color: var(--neutral-900);
    }
    
    div[data-testid="stHorizontalBlock"] .stRadio > div > label[data-checked="true"] {
        background: var(--ups-brown-700) !important;
        color: var(--ups-gold-500) !important;
        box-shadow: var(--shadow-md);
        border-color: var(--ups-brown-600);
    }
    
    /* ============================================
       BUTTONS - ENTERPRISE STYLE
       ============================================ */
    .stButton > button {
        background: var(--ups-brown-700);
        color: var(--ups-gold-500);
        font-weight: 600;
        font-size: 13px;
        border: none;
        padding: var(--space-3) var(--space-5);
        border-radius: var(--radius-md);
        transition: all 0.15s ease;
        box-shadow: var(--shadow-sm);
        letter-spacing: 0.01em;
    }
    
    .stButton > button:hover {
        background: var(--ups-brown-600);
        box-shadow: var(--shadow-md);
        transform: translateY(-1px);
    }
    
    .stButton > button:active {
        transform: translateY(0);
        box-shadow: var(--shadow-sm);
    }
    
    .stDownloadButton > button {
        background: white;
        color: var(--ups-brown-700);
        font-weight: 600;
        font-size: 13px;
        border: 1px solid var(--neutral-300);
        border-radius: var(--radius-md);
        transition: all 0.15s ease;
    }
    
    .stDownloadButton > button:hover {
        background: var(--neutral-50);
        border-color: var(--ups-brown-700);
        color: var(--ups-brown-700);
    }
    
    /* ============================================
       FORM CONTROLS
       ============================================ */
    .stSelectbox > div > div {
        border: 1px solid var(--neutral-300);
        border-radius: var(--radius-md);
        background: white;
        transition: all 0.15s ease;
    }
    
    .stSelectbox > div > div:hover {
        border-color: var(--neutral-400);
    }
    
    .stSelectbox > div > div:focus-within {
        border-color: var(--ups-gold-500);
        box-shadow: 0 0 0 3px rgba(255, 181, 0, 0.12);
    }
    
    .stDateInput > div > div > input {
        border: 1px solid var(--neutral-300);
        border-radius: var(--radius-md);
        padding: var(--space-2) var(--space-3);
        font-size: 13px;
    }
    
    .stDateInput > div > div:focus-within {
        border-color: var(--ups-gold-500);
        box-shadow: 0 0 0 3px rgba(255, 181, 0, 0.12);
    }
    
    .stCheckbox > label {
        font-size: 13px;
        font-weight: 500;
        color: var(--neutral-800);
    }
    
    /* ============================================
       CONTENT PANELS
       ============================================ */
    .content-panel {
        background: white;
        border: 1px solid var(--neutral-200);
        border-radius: var(--radius-lg);
        padding: var(--space-6);
        box-shadow: var(--shadow-sm);
    }
    
    .panel-header {
        display: flex;
        align-items: center;
        gap: var(--space-3);
        margin-bottom: var(--space-5);
        padding-bottom: var(--space-4);
        border-bottom: 1px solid var(--neutral-200);
    }
    
    .panel-icon {
        width: 32px;
        height: 32px;
        background: var(--neutral-100);
        border-radius: var(--radius-md);
        display: flex;
        align-items: center;
        justify-content: center;
        font-size: 16px;
    }
    
    .panel-title {
        font-size: 15px;
        font-weight: 600;
        color: var(--neutral-900);
    }
    
    .panel-description {
        font-size: 13px;
        color: var(--neutral-600);
    }
    
    /* ============================================
       ROUTE RESULTS CARDS
       ============================================ */
    .route-result-card {
        background: white;
        border: 1px solid var(--neutral-200);
        border-radius: var(--radius-lg);
        padding: var(--space-5);
        margin-bottom: var(--space-4);
        transition: all 0.15s ease;
    }
    
    .route-result-card:hover {
        border-color: var(--neutral-300);
        box-shadow: var(--shadow-md);
    }
    
    .route-result-card.recommended {
        border-left: 3px solid var(--success-500);
    }
    
    .route-tag {
        display: inline-flex;
        align-items: center;
        padding: var(--space-1) var(--space-2);
        background: var(--neutral-100);
        border-radius: var(--radius-sm);
        font-size: 11px;
        font-weight: 600;
        text-transform: uppercase;
        letter-spacing: 0.03em;
        color: var(--neutral-700);
    }
    
    .route-tag.success {
        background: var(--success-50);
        color: var(--success-600);
    }
    
    /* ============================================
       ALERTS & NOTIFICATIONS
       ============================================ */
    .stSuccess > div {
        background: var(--success-50) !important;
        border: 1px solid rgba(0, 135, 90, 0.2) !important;
        border-left: 3px solid var(--success-500) !important;
        border-radius: 0 var(--radius-md) var(--radius-md) 0 !important;
        padding: var(--space-3) var(--space-4) !important;
    }
    
    .stWarning > div {
        background: var(--warning-50) !important;
        border: 1px solid rgba(255, 153, 31, 0.2) !important;
        border-left: 3px solid var(--warning-500) !important;
        border-radius: 0 var(--radius-md) var(--radius-md) 0 !important;
        padding: var(--space-3) var(--space-4) !important;
    }
    
    .stInfo > div {
        background: var(--info-50) !important;
        border: 1px solid rgba(0, 101, 255, 0.2) !important;
        border-left: 3px solid var(--info-500) !important;
        border-radius: 0 var(--radius-md) var(--radius-md) 0 !important;
        padding: var(--space-3) var(--space-4) !important;
    }
    
    .stError > div {
        background: var(--error-50) !important;
        border: 1px solid rgba(222, 53, 11, 0.2) !important;
        border-left: 3px solid var(--error-500) !important;
        border-radius: 0 var(--radius-md) var(--radius-md) 0 !important;
        padding: var(--space-3) var(--space-4) !important;
    }
    
    /* ============================================
       EXPANDERS - COLLAPSIBLE SECTIONS
       ============================================ */
    .streamlit-expanderHeader {
        background: var(--neutral-50) !important;
        border: 1px solid var(--neutral-200) !important;
        border-radius: var(--radius-md) !important;
        font-size: 13px !important;
        font-weight: 600 !important;
        padding: var(--space-3) var(--space-4) !important;
        color: var(--neutral-800) !important;
    }
    
    .streamlit-expanderHeader:hover {
        background: var(--neutral-100) !important;
    }
    
    details[open] .streamlit-expanderHeader {
        border-radius: var(--radius-md) var(--radius-md) 0 0 !important;
        border-bottom: none !important;
    }
    
    .streamlit-expanderContent {
        border: 1px solid var(--neutral-200) !important;
        border-top: none !important;
        border-radius: 0 0 var(--radius-md) var(--radius-md) !important;
        padding: var(--space-4) !important;
        background: white !important;
    }
    
    /* ============================================
       DOWNLOAD SECTION
       ============================================ */
    .download-panel {
        background: var(--neutral-50);
        border: 1px solid var(--neutral-200);
        border-radius: var(--radius-lg);
        padding: var(--space-5);
        margin-top: var(--space-6);
    }
    
    .download-panel-header {
        font-size: 14px;
        font-weight: 600;
        color: var(--neutral-900);
        margin-bottom: var(--space-1);
    }
    
    .download-panel-desc {
        font-size: 13px;
        color: var(--neutral-600);
        margin-bottom: var(--space-4);
    }
    
    /* ============================================
       DIVIDERS
       ============================================ */
    hr {
        border: none;
        height: 1px;
        background: var(--neutral-200);
        margin: var(--space-6) 0;
    }
    
    /* ============================================
       SCROLLBAR
       ============================================ */
    ::-webkit-scrollbar {
        width: 8px;
        height: 8px;
    }
    
    ::-webkit-scrollbar-track {
        background: var(--neutral-100);
    }
    
    ::-webkit-scrollbar-thumb {
        background: var(--neutral-300);
        border-radius: 4px;
    }
    
    ::-webkit-scrollbar-thumb:hover {
        background: var(--neutral-400);
    }
    
    /* ============================================
       UTILITY CLASSES
       ============================================ */
    .text-muted { color: var(--neutral-600); }
    .text-small { font-size: 12px; }
    .text-xs { font-size: 11px; }
    .font-medium { font-weight: 500; }
    .font-semibold { font-weight: 600; }
    .uppercase { text-transform: uppercase; letter-spacing: 0.04em; }
    
    .mt-1 { margin-top: var(--space-1); }
    .mt-2 { margin-top: var(--space-2); }
    .mt-3 { margin-top: var(--space-3); }
    .mt-4 { margin-top: var(--space-4); }
    .mt-6 { margin-top: var(--space-6); }
    
    .mb-1 { margin-bottom: var(--space-1); }
    .mb-2 { margin-bottom: var(--space-2); }
    .mb-3 { margin-bottom: var(--space-3); }
    .mb-4 { margin-bottom: var(--space-4); }
    .mb-6 { margin-bottom: var(--space-6); }
    
    /* ============================================
       ANIMATIONS
       ============================================ */
    @keyframes fadeIn {
        from { opacity: 0; transform: translateY(4px); }
        to { opacity: 1; transform: translateY(0); }
    }
    
    .kpi-card, .content-panel, .route-result-card {
        animation: fadeIn 0.3s ease-out;
    }
    
    /* ============================================
       RESPONSIVE ADJUSTMENTS
       ============================================ */
    @media (max-width: 768px) {
        .app-header-inner {
            flex-direction: column;
            align-items: flex-start;
            gap: var(--space-3);
        }
        
        .app-header-meta {
            width: 100%;
            justify-content: space-between;
        }
    }
    </style>
""", unsafe_allow_html=True)

# ============================================================================
# HEADER - Enterprise Branding
# ============================================================================
st.markdown("""
<div class="app-header">
    <div class="app-header-inner">
        <div class="app-header-logo">
            <img src="https://upload.wikimedia.org/wikipedia/commons/6/6b/United_Parcel_Service_logo_2014.svg" 
                 alt="UPS Healthcare Logistics">
        </div>
        <div class="app-header-content">
            <h1 class="app-header-title">Flight Routing System</h1>
            <p class="app-header-subtitle">Healthcare Logistics · Global Route Optimization</p>
        </div>
        <div class="app-header-meta">
            <div class="status-indicator">Operational</div>
            <div class="marken-badge">Powered by Marken</div>
        </div>
    </div>
</div>
""", unsafe_allow_html=True)

# System quick reference - clean enterprise style
st.markdown("""
<div style="display: flex; gap: 16px; margin-bottom: 24px;">
    <div style="flex: 1; background: white; border: 1px solid #e8eaed; border-radius: 8px; padding: 16px; border-left: 3px solid #FFB500;">
        <div style="font-size: 12px; font-weight: 600; color: #505f79; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 4px;">System Purpose</div>
        <div style="font-size: 13px; color: #172b4d;">Global flight routing optimization for healthcare logistics shipments</div>
    </div>
    <div style="flex: 1; background: white; border: 1px solid #e8eaed; border-radius: 8px; padding: 16px; border-left: 3px solid #FFB500;">
        <div style="font-size: 12px; font-weight: 600; color: #505f79; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 4px;">Connection Windows</div>
        <div style="font-size: 13px; color: #172b4d;">Minimum: 60 minutes · Maximum: 24 hours</div>
    </div>
    <div style="flex: 1; background: white; border: 1px solid #e8eaed; border-radius: 8px; padding: 16px; border-left: 3px solid #FFB500;">
        <div style="font-size: 12px; font-weight: 600; color: #505f79; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 4px;">Available Modules</div>
        <div style="font-size: 13px; color: #172b4d;">Tracked Routes · Custom Routes · RadioPharma</div>
    </div>
</div>
""", unsafe_allow_html=True)

with st.expander("System Documentation", expanded=False):
    st.markdown("""
    ### Overview
    This dashboard provides UPS Healthcare Logistics with an automated flight routing system for optimizing 
    package shipments between global airports. The system analyzes flight combinations to identify the most 
    efficient routing options based on time constraints and operational requirements.
    
    ### Route Prioritization
    
    | Priority | Method | Use Case |
    |:--------:|--------|----------|
    | 1 | **Fewest Stops** | Sensitive shipments requiring minimal handling |
    | 2 | **Fastest Arrival** | Time-critical deliveries |
    | 3 | **Direct Flights** | Non-stop options when available |
    | 4 | **Date Extension** | Auto-search up to 7 days forward |
    
    ### Connection Requirements
    - **Minimum Connection**: 60 minutes between arrival and departure
    - **Maximum Connection**: 24 hours to prevent storage issues
    - **Complex Routes**: Routes with 5+ stops require logistics team consultation
    
    ### Results Interpretation
    
    **Fewest Stops** (Primary recommendation)
    - Minimizes cargo handling and transfer risk
    - Sorted by total journey duration
    
    **Fastest Arriving**
    - Optimizes for earliest destination arrival
    - May include additional connections
    """)

st.markdown("---")

@st.cache_data
def load_data(file):
    """Load and parse the Excel file"""
    try:
        # Read main sheets
        schedule_df = pd.read_excel(file, sheet_name='SchedDateLocalTimeFlightSchedul')
        routes_df = pd.read_excel(file, sheet_name='Data')
        
        # Convert date columns
        schedule_df['Start Date (LZ)'] = pd.to_datetime(schedule_df['Start Date (LZ)'], errors='coerce')
        schedule_df['End Date (LZ)'] = pd.to_datetime(schedule_df['End Date (LZ)'], errors='coerce')
        
        # Ensure string format for time columns
        schedule_df['Sched Out(L)'] = schedule_df['Sched Out(L)'].astype(str)
        schedule_df['Sched In(L)'] = schedule_df['Sched In(L)'].astype(str)
        schedule_df['Blkhr'] = schedule_df['Blkhr'].astype(str)
        schedule_df['DOW(S)'] = schedule_df['DOW(S)'].astype(str)
        
        # Try to read RadioPharma Info sheet
        rp_config = None
        try:
            rp_df = pd.read_excel(file, sheet_name='RP Info')
            
            # Extract data from RP Info sheet
            # Column 0: Prohibited flights (skip header row)
            prohibited_flights = set(rp_df.iloc[1:, 0].dropna().astype(str).str.strip().tolist())
            
            # Column 3: Approved Origins (skip header row)
            approved_origins = set(rp_df.iloc[1:, 3].dropna().astype(str).str.strip().tolist())
            
            # Column 4: Approved Destinations (skip header row)
            approved_destinations = set(rp_df.iloc[1:, 4].dropna().astype(str).str.strip().tolist())
            
            # Transit points are the same as approved destinations
            approved_transit = approved_destinations.copy()
            
            rp_config = {
                'prohibited_flights': prohibited_flights,
                'approved_origins': approved_origins,
                'approved_destinations': approved_destinations,
                'approved_transit': approved_transit
            }
        except Exception as rp_error:
            # RP Info sheet not found or error reading it - RadioPharma tab will be disabled
            rp_config = None
        
        return schedule_df, routes_df, rp_config
    except Exception as e:
        st.error(f"Error loading data: {str(e)}")
        return None, None, None

def is_flight_available_on_date(dow_string, date):
    """Check if flight operates on given date based on DOW(S) string"""
    if pd.isna(dow_string) or dow_string == 'nan' or dow_string == '':
        return False
    
    dow_string = str(dow_string).strip()
    weekday = date.weekday()  # Python: 0=Monday, 6=Sunday
    
    if weekday < len(dow_string):
        return dow_string[weekday] != '.'
    return False

def parse_time_to_minutes(time_str):
    """Convert time string (HH:MM) to minutes from midnight"""
    try:
        if pd.isna(time_str) or str(time_str) == 'nan':
            return None
        time_str = str(time_str).strip()
        if ':' in time_str:
            parts = time_str.split(':')
            hours = int(parts[0])
            minutes = int(parts[1]) if len(parts) > 1 else 0
            return hours * 60 + minutes
        return None
    except:
        return None

def format_duration(minutes):
    """Format duration in minutes to readable format"""
    if minutes is None:
        return "N/A"
    hours = minutes // 60
    mins = minutes % 60
    return f"{hours}h {mins}m"

def find_direct_flights(schedule_df, origin, destination, date, days_ahead=7):
    """Find direct flights - prioritize same day, then search up to 7 days if needed"""
    try:
        route_flights = schedule_df[
            (schedule_df['Orig'] == origin) & 
            (schedule_df['Dest'] == destination)
        ].copy()
        
        if route_flights.empty:
            return []
        
        results = []
        all_flights_with_arrival = []
        
        for day_offset in range(0, 4):
            check_date = date + timedelta(days=day_offset)
            flights_on_date = []
            
            for idx, flight in route_flights.iterrows():
                try:
                    if is_flight_available_on_date(flight['DOW(S)'], check_date):
                        if pd.notna(flight['Start Date (LZ)']) and pd.notna(flight['End Date (LZ)']):
                            if flight['Start Date (LZ)'].date() <= check_date.date() <= flight['End Date (LZ)'].date():
                                flight_copy = flight.copy()
                                flight_copy['flight_date'] = check_date
                                dep_time = parse_time_to_minutes(flight['Sched Out(L)'])
                                arr_time = parse_time_to_minutes(flight['Sched In(L)'])
                                flight_copy['dep_minutes'] = dep_time if dep_time else 0
                                flight_copy['arr_minutes'] = arr_time if arr_time else 0
                                
                                arrival_date = check_date
                                if arr_time and dep_time and arr_time < dep_time:
                                    arrival_date = check_date + timedelta(days=1)
                                
                                if arr_time is not None:
                                    flight_copy['arrival_datetime'] = arrival_date.replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(minutes=arr_time)
                                else:
                                    flight_copy['arrival_datetime'] = arrival_date
                                
                                flight_copy['day_offset'] = day_offset
                                flights_on_date.append(flight_copy)
                                all_flights_with_arrival.append(flight_copy)
                except:
                    continue
            
            if flights_on_date and day_offset == 0:
                flights_on_date.sort(key=lambda x: x['arrival_datetime'])
                results.append({
                    'date': check_date,
                    'flights': flights_on_date,
                    'days_from_requested': day_offset,
                    'arrives_earlier': False
                })
        
        if results and len(results) > 0:
            same_day_flights = results[0]['flights']
            earliest_same_day_arrival = min(f['arrival_datetime'] for f in same_day_flights)
            
            for day_offset in range(1, 4):
                check_date = date + timedelta(days=day_offset)
                earlier_flights = [f for f in all_flights_with_arrival 
                                  if f['day_offset'] == day_offset 
                                  and f['arrival_datetime'] < earliest_same_day_arrival]
                
                if earlier_flights:
                    earlier_flights.sort(key=lambda x: x['arrival_datetime'])
                    results.append({
                        'date': check_date,
                        'flights': earlier_flights,
                        'days_from_requested': day_offset,
                        'arrives_earlier': True
                    })
        
        if not results:
            for day_offset in range(1, days_ahead + 1):
                check_date = date + timedelta(days=day_offset)
                flights_on_date = []
                
                for idx, flight in route_flights.iterrows():
                    try:
                        if is_flight_available_on_date(flight['DOW(S)'], check_date):
                            if pd.notna(flight['Start Date (LZ)']) and pd.notna(flight['End Date (LZ)']):
                                if flight['Start Date (LZ)'].date() <= check_date.date() <= flight['End Date (LZ)'].date():
                                    flight_copy = flight.copy()
                                    flight_copy['flight_date'] = check_date
                                    dep_time = parse_time_to_minutes(flight['Sched Out(L)'])
                                    arr_time = parse_time_to_minutes(flight['Sched In(L)'])
                                    flight_copy['dep_minutes'] = dep_time if dep_time else 0
                                    flight_copy['arr_minutes'] = arr_time if arr_time else 0
                                    
                                    arrival_date = check_date
                                    if arr_time and dep_time and arr_time < dep_time:
                                        arrival_date = check_date + timedelta(days=1)
                                    
                                    if arr_time is not None:
                                        flight_copy['arrival_datetime'] = arrival_date.replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(minutes=arr_time)
                                    else:
                                        flight_copy['arrival_datetime'] = arrival_date
                                    
                                    flights_on_date.append(flight_copy)
                    except:
                        continue
                
                if flights_on_date:
                    flights_on_date.sort(key=lambda x: x['arrival_datetime'])
                    results.append({
                        'date': check_date,
                        'flights': flights_on_date,
                        'days_from_requested': day_offset,
                        'arrives_earlier': False
                    })
                    
                    if len(results) >= 3:
                        break
        
        return results
        
    except Exception as e:
        return []

def build_network(schedule_df, start_date, days_ahead=7):
    """Build flight network for routing with proper date/time logic - OPTIMIZED"""
    network = {}
    
    try:
        for day_offset in range(days_ahead + 1):
            check_date = start_date + timedelta(days=day_offset)
            
            for idx, flight in schedule_df.iterrows():
                try:
                    if is_flight_available_on_date(flight['DOW(S)'], check_date):
                        if pd.notna(flight['Start Date (LZ)']) and pd.notna(flight['End Date (LZ)']):
                            if flight['Start Date (LZ)'].date() <= check_date.date() <= flight['End Date (LZ)'].date():
                                origin = str(flight['Orig'])
                                dest = str(flight['Dest'])
                                
                                if origin not in network:
                                    network[origin] = []
                                
                                dep_time = parse_time_to_minutes(flight['Sched Out(L)'])
                                arr_time = parse_time_to_minutes(flight['Sched In(L)'])
                                
                                blkhr_str = str(flight['Blkhr'])
                                flight_duration = 0
                                if pd.notna(flight['Blkhr']) and blkhr_str != 'nan':
                                    if ':' in blkhr_str:
                                        parts = blkhr_str.split(':')
                                        hours = int(parts[0]) if parts[0] else 0
                                        minutes = int(parts[1]) if len(parts) > 1 else 0
                                        flight_duration = hours * 60 + minutes
                                
                                if dep_time is not None and arr_time is not None and flight_duration > 0:
                                    arrival_date = check_date
                                    if arr_time < dep_time:
                                        arrival_date = check_date + timedelta(days=1)
                                    
                                    # Pre-calculate arrival_datetime for faster sorting later
                                    arrival_datetime = arrival_date.replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(minutes=arr_time)
                                    
                                    network[origin].append({
                                        'destination': dest,
                                        'departure': dep_time,
                                        'arrival': arr_time,
                                        'arrival_date': arrival_date,
                                        'arrival_datetime': arrival_datetime,
                                        'dep_str': str(flight['Sched Out(L)']),
                                        'arr_str': str(flight['Sched In(L)']),
                                        'duration_str': str(flight['Blkhr']),
                                        'carrier': str(flight.get('Carrier', 'N/A')),
                                        'flight_num': f"{flight.get('Carrier', '')}{flight.get('Flight #', '')}",
                                        'duration': flight_duration,
                                        'date': check_date,
                                        'day_offset': day_offset
                                    })
                except:
                    continue
    except:
        pass
    
    # Sort flights by arrival time for each origin (helps with pruning later)
    for origin in network:
        network[origin].sort(key=lambda x: (x['date'], x['arrival_datetime']))
    
    return network

def build_radiopharma_network(schedule_df, start_date, rp_config, days_ahead=7):
    """Build flight network for RadioPharma routing - excludes prohibited flights and non-approved airports"""
    network = {}
    
    # Extract config
    prohibited_flights = rp_config['prohibited_flights']
    approved_transit = rp_config['approved_transit']
    
    try:
        for day_offset in range(days_ahead + 1):
            check_date = start_date + timedelta(days=day_offset)
            
            for idx, flight in schedule_df.iterrows():
                try:
                    # Get flight number and check if prohibited
                    flight_num = str(flight.get('Flight #', ''))
                    if flight_num in prohibited_flights:
                        continue  # Skip prohibited flights
                    
                    if is_flight_available_on_date(flight['DOW(S)'], check_date):
                        if pd.notna(flight['Start Date (LZ)']) and pd.notna(flight['End Date (LZ)']):
                            if flight['Start Date (LZ)'].date() <= check_date.date() <= flight['End Date (LZ)'].date():
                                origin = str(flight['Orig'])
                                dest = str(flight['Dest'])
                                
                                # Only include flights where destination is an approved transit/destination point
                                if dest not in approved_transit:
                                    continue
                                
                                if origin not in network:
                                    network[origin] = []
                                
                                dep_time = parse_time_to_minutes(flight['Sched Out(L)'])
                                arr_time = parse_time_to_minutes(flight['Sched In(L)'])
                                
                                blkhr_str = str(flight['Blkhr'])
                                flight_duration = 0
                                if pd.notna(flight['Blkhr']) and blkhr_str != 'nan':
                                    if ':' in blkhr_str:
                                        parts = blkhr_str.split(':')
                                        hours = int(parts[0]) if parts[0] else 0
                                        minutes = int(parts[1]) if len(parts) > 1 else 0
                                        flight_duration = hours * 60 + minutes
                                
                                if dep_time is not None and arr_time is not None and flight_duration > 0:
                                    arrival_date = check_date
                                    if arr_time < dep_time:
                                        arrival_date = check_date + timedelta(days=1)
                                    
                                    arrival_datetime = arrival_date.replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(minutes=arr_time)
                                    
                                    network[origin].append({
                                        'destination': dest,
                                        'departure': dep_time,
                                        'arrival': arr_time,
                                        'arrival_date': arrival_date,
                                        'arrival_datetime': arrival_datetime,
                                        'dep_str': str(flight['Sched Out(L)']),
                                        'arr_str': str(flight['Sched In(L)']),
                                        'duration_str': str(flight['Blkhr']),
                                        'carrier': str(flight.get('Carrier', 'N/A')),
                                        'flight_num': f"{flight.get('Carrier', '')}{flight.get('Flight #', '')}",
                                        'duration': flight_duration,
                                        'date': check_date,
                                        'day_offset': day_offset
                                    })
                except:
                    continue
    except:
        pass
    
    # Sort flights by arrival time for each origin
    for origin in network:
        network[origin].sort(key=lambda x: (x['date'], x['arrival_datetime']))
    
    return network

def find_radiopharma_routes_for_date(network, origin, destination, target_date, rp_config, max_stops=10, min_departure_time=0):
    """
    Find RadioPharma routes that DEPART on the target_date.
    Only allows transit through approved transit points.
    """
    # Extract config
    approved_origins = rp_config['approved_origins']
    approved_destinations = rp_config['approved_destinations']
    approved_transit = rp_config['approved_transit']
    
    if origin not in network:
        return []
    
    if origin not in approved_origins:
        return []
    
    if destination not in approved_destinations:
        return []
    
    all_routes = []
    
    # Get ONLY flights from origin on the TARGET DATE that depart at or after min_departure_time
    initial_flights = [f for f in network.get(origin, []) 
                      if f['date'].date() == target_date.date() and f['departure'] >= min_departure_time]
    
    if not initial_flights:
        return []
    
    # ================================================================
    # SEARCH 1: Find routes with FEWEST STOPS (level-by-level BFS)
    # ================================================================
    def find_fewest_stops_routes():
        routes_found = []
        min_stops_found = None
        
        # ============ LEVEL 0: Direct flights (0 stops) ============
        for first_flight in initial_flights:
            if first_flight['destination'] == destination:
                first_arrival_date = first_flight.get('arrival_date', first_flight['date'])
                first_arrival_time = first_flight['arrival']
                
                if first_arrival_time is not None:
                    arrival_datetime = first_arrival_date.replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(minutes=first_arrival_time)
                else:
                    arrival_datetime = first_arrival_date
                
                routes_found.append({
                    'path': [origin, destination],
                    'stops': 0,
                    'total_duration': first_flight['duration'],
                    'route_info': [{
                        'from': origin,
                        'to': destination,
                        'date': first_flight['date'],
                        'departure': first_flight['dep_str'],
                        'arrival': first_flight['arr_str'],
                        'duration': first_flight['duration'],
                        'duration_str': first_flight['duration_str'],
                        'carrier': first_flight['carrier'],
                        'flight': first_flight['flight_num'],
                        'wait_time': 0
                    }],
                    'start_date': first_flight['date'],
                    'end_date': first_arrival_date,
                    'arrival_datetime': arrival_datetime
                })
                min_stops_found = 0
        
        if min_stops_found == 0:
            return routes_found
        
        from collections import deque
        queue = deque()
        
        for first_flight in initial_flights:
            if first_flight['destination'] == destination:
                continue
            
            # Only continue if destination is approved transit point
            if first_flight['destination'] not in approved_transit:
                continue
            
            first_arrival_date = first_flight.get('arrival_date', first_flight['date'])
            first_arrival_time = first_flight['arrival']
            
            if first_arrival_time is not None:
                first_arrival_datetime = first_arrival_date.replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(minutes=first_arrival_time)
            else:
                first_arrival_datetime = first_arrival_date
            
            queue.append({
                'current': first_flight['destination'],
                'path': [origin, first_flight['destination']],
                'arrival_time': first_arrival_time,
                'arrival_date': first_arrival_date,
                'arrival_datetime': first_arrival_datetime,
                'total_duration': first_flight['duration'],
                'route_info': [{
                    'from': origin,
                    'to': first_flight['destination'],
                    'date': first_flight['date'],
                    'departure': first_flight['dep_str'],
                    'arrival': first_flight['arr_str'],
                    'duration': first_flight['duration'],
                    'duration_str': first_flight['duration_str'],
                    'carrier': first_flight['carrier'],
                    'flight': first_flight['flight_num'],
                    'wait_time': 0
                }],
                'stops': 0
            })
        
        visited = set()
        
        while queue:
            state = queue.popleft()
            
            current = state['current']
            path = state['path']
            current_stops = state['stops']
            
            if min_stops_found is not None and current_stops >= min_stops_found:
                continue
            
            if current_stops >= max_stops:
                continue
            
            state_key = (current, tuple(path))
            if state_key in visited:
                continue
            visited.add(state_key)
            
            last_arrival_time = state['arrival_time']
            last_arrival_date = state['arrival_date']
            
            if current not in network:
                continue
            
            for next_flight in network[current]:
                if next_flight['destination'] in path:
                    continue
                
                # Only allow transit through approved points (unless it's final destination)
                if next_flight['destination'] != destination and next_flight['destination'] not in approved_transit:
                    continue
                
                if next_flight['date'] < last_arrival_date:
                    continue
                if next_flight['date'] > last_arrival_date + timedelta(days=3):
                    continue
                
                min_connection = 60
                
                if next_flight['date'].date() > last_arrival_date.date():
                    days_diff = (next_flight['date'].date() - last_arrival_date.date()).days
                    wait_time = (1440 - last_arrival_time) + ((days_diff - 1) * 1440) + next_flight['departure']
                elif next_flight['date'].date() == last_arrival_date.date():
                    if next_flight['departure'] >= last_arrival_time + min_connection:
                        wait_time = next_flight['departure'] - last_arrival_time
                    else:
                        continue
                else:
                    continue
                
                if wait_time < min_connection or wait_time > 1440:
                    continue
                
                next_arrival_date = next_flight.get('arrival_date', next_flight['date'])
                next_arrival_time = next_flight['arrival']
                
                if next_arrival_time is not None:
                    next_arrival_datetime = next_arrival_date.replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(minutes=next_arrival_time)
                else:
                    next_arrival_datetime = next_arrival_date
                
                new_route_info = state['route_info'] + [{
                    'from': current,
                    'to': next_flight['destination'],
                    'date': next_flight['date'],
                    'departure': next_flight['dep_str'],
                    'arrival': next_flight['arr_str'],
                    'duration': next_flight['duration'],
                    'duration_str': next_flight['duration_str'],
                    'carrier': next_flight['carrier'],
                    'flight': next_flight['flight_num'],
                    'wait_time': wait_time
                }]
                
                new_path = path + [next_flight['destination']]
                new_stops = current_stops + 1
                new_duration = state['total_duration'] + wait_time + next_flight['duration']
                
                if next_flight['destination'] == destination:
                    if min_stops_found is None or new_stops <= min_stops_found:
                        min_stops_found = new_stops
                        routes_found.append({
                            'path': new_path,
                            'stops': new_stops,
                            'total_duration': new_duration,
                            'route_info': new_route_info,
                            'start_date': new_route_info[0]['date'],
                            'end_date': next_arrival_date,
                            'arrival_datetime': next_arrival_datetime
                        })
                else:
                    queue.append({
                        'current': next_flight['destination'],
                        'path': new_path,
                        'arrival_time': next_arrival_time,
                        'arrival_date': next_arrival_date,
                        'arrival_datetime': next_arrival_datetime,
                        'total_duration': new_duration,
                        'route_info': new_route_info,
                        'stops': new_stops
                    })
        
        if routes_found:
            actual_min = min(r['stops'] for r in routes_found)
            routes_found = [r for r in routes_found if r['stops'] == actual_min]
        
        return routes_found
    
    # ================================================================
    # SEARCH 2: Find FASTEST ARRIVING routes
    # ================================================================
    def find_fastest_routes():
        routes_found = []
        
        from collections import deque
        queue = deque()
        
        for first_flight in initial_flights:
            first_arrival_date = first_flight.get('arrival_date', first_flight['date'])
            first_arrival_time = first_flight['arrival']
            
            if first_arrival_time is not None:
                first_arrival_datetime = first_arrival_date.replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(minutes=first_arrival_time)
            else:
                first_arrival_datetime = first_arrival_date
            
            if first_flight['destination'] == destination:
                routes_found.append({
                    'path': [origin, destination],
                    'stops': 0,
                    'total_duration': first_flight['duration'],
                    'route_info': [{
                        'from': origin,
                        'to': destination,
                        'date': first_flight['date'],
                        'departure': first_flight['dep_str'],
                        'arrival': first_flight['arr_str'],
                        'duration': first_flight['duration'],
                        'duration_str': first_flight['duration_str'],
                        'carrier': first_flight['carrier'],
                        'flight': first_flight['flight_num'],
                        'wait_time': 0
                    }],
                    'start_date': first_flight['date'],
                    'end_date': first_arrival_date,
                    'arrival_datetime': first_arrival_datetime
                })
            elif first_flight['destination'] in approved_transit:
                queue.append({
                    'current': first_flight['destination'],
                    'path': [origin, first_flight['destination']],
                    'arrival_time': first_arrival_time,
                    'arrival_date': first_arrival_date,
                    'arrival_datetime': first_arrival_datetime,
                    'total_duration': first_flight['duration'],
                    'route_info': [{
                        'from': origin,
                        'to': first_flight['destination'],
                        'date': first_flight['date'],
                        'departure': first_flight['dep_str'],
                        'arrival': first_flight['arr_str'],
                        'duration': first_flight['duration'],
                        'duration_str': first_flight['duration_str'],
                        'carrier': first_flight['carrier'],
                        'flight': first_flight['flight_num'],
                        'wait_time': 0
                    }]
                })
        
        visited = set()
        iterations = 0
        max_iterations = 100000
        
        while queue and iterations < max_iterations:
            iterations += 1
            state = queue.popleft()
            
            current = state['current']
            path = state['path']
            
            if len(path) - 1 >= max_stops + 1:
                continue
            
            state_key = (current, tuple(path))
            if state_key in visited:
                continue
            visited.add(state_key)
            
            last_arrival_time = state['arrival_time']
            last_arrival_date = state['arrival_date']
            
            if current not in network:
                continue
            
            for next_flight in network[current]:
                if next_flight['destination'] in path:
                    continue
                
                # Only allow transit through approved points
                if next_flight['destination'] != destination and next_flight['destination'] not in approved_transit:
                    continue
                
                if next_flight['date'] < last_arrival_date:
                    continue
                if next_flight['date'] > last_arrival_date + timedelta(days=3):
                    continue
                
                min_connection = 60
                
                if next_flight['date'].date() > last_arrival_date.date():
                    days_diff = (next_flight['date'].date() - last_arrival_date.date()).days
                    wait_time = (1440 - last_arrival_time) + ((days_diff - 1) * 1440) + next_flight['departure']
                elif next_flight['date'].date() == last_arrival_date.date():
                    if next_flight['departure'] >= last_arrival_time + min_connection:
                        wait_time = next_flight['departure'] - last_arrival_time
                    else:
                        continue
                else:
                    continue
                
                if wait_time < min_connection or wait_time > 1440:
                    continue
                
                next_arrival_date = next_flight.get('arrival_date', next_flight['date'])
                next_arrival_time = next_flight['arrival']
                
                if next_arrival_time is not None:
                    next_arrival_datetime = next_arrival_date.replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(minutes=next_arrival_time)
                else:
                    next_arrival_datetime = next_arrival_date
                
                new_route_info = state['route_info'] + [{
                    'from': current,
                    'to': next_flight['destination'],
                    'date': next_flight['date'],
                    'departure': next_flight['dep_str'],
                    'arrival': next_flight['arr_str'],
                    'duration': next_flight['duration'],
                    'duration_str': next_flight['duration_str'],
                    'carrier': next_flight['carrier'],
                    'flight': next_flight['flight_num'],
                    'wait_time': wait_time
                }]
                
                new_path = path + [next_flight['destination']]
                new_duration = state['total_duration'] + wait_time + next_flight['duration']
                
                if next_flight['destination'] == destination:
                    routes_found.append({
                        'path': new_path,
                        'stops': len(new_path) - 2,
                        'total_duration': new_duration,
                        'route_info': new_route_info,
                        'start_date': new_route_info[0]['date'],
                        'end_date': next_arrival_date,
                        'arrival_datetime': next_arrival_datetime
                    })
                else:
                    queue.append({
                        'current': next_flight['destination'],
                        'path': new_path,
                        'arrival_time': next_arrival_time,
                        'arrival_date': next_arrival_date,
                        'arrival_datetime': next_arrival_datetime,
                        'total_duration': new_duration,
                        'route_info': new_route_info
                    })
        
        return routes_found
    
    # Run BOTH searches
    fewest_stops_routes = find_fewest_stops_routes()
    fastest_routes = find_fastest_routes()
    
    # Combine all routes
    all_routes = fewest_stops_routes + fastest_routes
    
    # Remove duplicates
    unique_routes = []
    seen_routes = set()
    
    for route in all_routes:
        route_id = tuple([
            (leg['from'], leg['to'], leg['date'].date(), leg['departure'])
            for leg in route['route_info']
        ])
        
        if route_id not in seen_routes:
            seen_routes.add(route_id)
            unique_routes.append(route)
    
    return unique_routes

def find_all_routes_for_date(network, origin, destination, target_date, max_stops=10, min_departure_time=0):
    """
    Find routes that DEPART on the target_date.
    
    Args:
        min_departure_time: Minimum departure time in minutes from midnight (e.g., 600 = 10:00 AM)
    
    FEWEST STOPS: Uses level-by-level BFS to GUARANTEE finding minimum stops.
    - First checks 0 stops (direct flights)
    - Then checks 1 stop (all CGN→X→BFI combinations)
    - Then checks 2 stops, etc.
    
    FASTEST ARRIVING: Explores ALL routes and returns those arriving earliest.
    """
    if origin not in network:
        return []
    
    all_routes = []
    
    # Get ONLY flights from origin on the TARGET DATE that depart at or after min_departure_time
    initial_flights = [f for f in network.get(origin, []) 
                      if f['date'].date() == target_date.date() and f['departure'] >= min_departure_time]
    
    if not initial_flights:
        return []
    
    # ================================================================
    # SEARCH 1: Find routes with FEWEST STOPS (level-by-level BFS)
    # This GUARANTEES finding the minimum number of stops
    # ================================================================
    def find_fewest_stops_routes():
        """
        Level-by-level search:
        Level 0: Direct flights (0 stops)
        Level 1: One intermediate stop
        Level 2: Two intermediate stops
        etc.
        """
        routes_found = []
        min_stops_found = None
        
        # ============ LEVEL 0: Direct flights (0 stops) ============
        for first_flight in initial_flights:
            if first_flight['destination'] == destination:
                first_arrival_date = first_flight.get('arrival_date', first_flight['date'])
                first_arrival_time = first_flight['arrival']
                
                if first_arrival_time is not None:
                    arrival_datetime = first_arrival_date.replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(minutes=first_arrival_time)
                else:
                    arrival_datetime = first_arrival_date
                
                routes_found.append({
                    'path': [origin, destination],
                    'stops': 0,
                    'total_duration': first_flight['duration'],
                    'route_info': [{
                        'from': origin,
                        'to': destination,
                        'date': first_flight['date'],
                        'departure': first_flight['dep_str'],
                        'arrival': first_flight['arr_str'],
                        'duration': first_flight['duration'],
                        'duration_str': first_flight['duration_str'],
                        'carrier': first_flight['carrier'],
                        'flight': first_flight['flight_num'],
                        'wait_time': 0
                    }],
                    'start_date': first_flight['date'],
                    'end_date': first_arrival_date,
                    'arrival_datetime': arrival_datetime
                })
                min_stops_found = 0
        
        # If we found direct flights, return them (0 stops is minimum)
        if min_stops_found == 0:
            return routes_found
        
        # ============ LEVEL 1+: Routes with stops ============
        # Use BFS exploring ALL connections at each level
        
        from collections import deque
        
        # Queue entries: (path, last_arrival_time, last_arrival_date, route_info, total_duration)
        queue = deque()
        
        # Initialize with all first flights (that don't go directly to destination)
        for first_flight in initial_flights:
            if first_flight['destination'] == destination:
                continue  # Already handled above
            
            first_arrival_date = first_flight.get('arrival_date', first_flight['date'])
            first_arrival_time = first_flight['arrival']
            
            if first_arrival_time is not None:
                first_arrival_datetime = first_arrival_date.replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(minutes=first_arrival_time)
            else:
                first_arrival_datetime = first_arrival_date
            
            queue.append({
                'current': first_flight['destination'],
                'path': [origin, first_flight['destination']],
                'arrival_time': first_arrival_time,
                'arrival_date': first_arrival_date,
                'arrival_datetime': first_arrival_datetime,
                'total_duration': first_flight['duration'],
                'route_info': [{
                    'from': origin,
                    'to': first_flight['destination'],
                    'date': first_flight['date'],
                    'departure': first_flight['dep_str'],
                    'arrival': first_flight['arr_str'],
                    'duration': first_flight['duration'],
                    'duration_str': first_flight['duration_str'],
                    'carrier': first_flight['carrier'],
                    'flight': first_flight['flight_num'],
                    'wait_time': 0
                }],
                'stops': 0  # Current intermediate stops (will be 1 when we reach dest)
            })
        
        visited = set()
        
        while queue:
            state = queue.popleft()
            
            current = state['current']
            path = state['path']
            current_stops = state['stops']
            
            # If we already found routes with fewer stops, skip
            if min_stops_found is not None and current_stops >= min_stops_found:
                continue
            
            # Skip if too many stops
            if current_stops >= max_stops:
                continue
            
            state_key = (current, tuple(path))
            if state_key in visited:
                continue
            visited.add(state_key)
            
            last_arrival_time = state['arrival_time']
            last_arrival_date = state['arrival_date']
            
            # Check ALL flights from current airport
            if current not in network:
                continue
            
            for next_flight in network[current]:
                # Skip if creates a cycle
                if next_flight['destination'] in path:
                    continue
                
                # Check connection timing
                if next_flight['date'] < last_arrival_date:
                    continue
                if next_flight['date'] > last_arrival_date + timedelta(days=3):
                    continue
                
                min_connection = 60  # 1 hour minimum
                
                # Calculate wait time
                if next_flight['date'].date() > last_arrival_date.date():
                    days_diff = (next_flight['date'].date() - last_arrival_date.date()).days
                    wait_time = (1440 - last_arrival_time) + ((days_diff - 1) * 1440) + next_flight['departure']
                elif next_flight['date'].date() == last_arrival_date.date():
                    if next_flight['departure'] >= last_arrival_time + min_connection:
                        wait_time = next_flight['departure'] - last_arrival_time
                    else:
                        continue  # Not enough connection time
                else:
                    continue
                
                if wait_time < min_connection or wait_time > 1440:
                    continue
                
                # Valid connection found!
                next_arrival_date = next_flight.get('arrival_date', next_flight['date'])
                next_arrival_time = next_flight['arrival']
                
                if next_arrival_time is not None:
                    next_arrival_datetime = next_arrival_date.replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(minutes=next_arrival_time)
                else:
                    next_arrival_datetime = next_arrival_date
                
                new_route_info = state['route_info'] + [{
                    'from': current,
                    'to': next_flight['destination'],
                    'date': next_flight['date'],
                    'departure': next_flight['dep_str'],
                    'arrival': next_flight['arr_str'],
                    'duration': next_flight['duration'],
                    'duration_str': next_flight['duration_str'],
                    'carrier': next_flight['carrier'],
                    'flight': next_flight['flight_num'],
                    'wait_time': wait_time
                }]
                
                new_path = path + [next_flight['destination']]
                new_stops = current_stops + 1
                new_duration = state['total_duration'] + wait_time + next_flight['duration']
                
                # Check if we reached destination
                if next_flight['destination'] == destination:
                    # Found a route!
                    if min_stops_found is None or new_stops <= min_stops_found:
                        min_stops_found = new_stops
                        routes_found.append({
                            'path': new_path,
                            'stops': new_stops,
                            'total_duration': new_duration,
                            'route_info': new_route_info,
                            'start_date': new_route_info[0]['date'],
                            'end_date': next_arrival_date,
                            'arrival_datetime': next_arrival_datetime
                        })
                else:
                    # Continue searching - add to queue
                    queue.append({
                        'current': next_flight['destination'],
                        'path': new_path,
                        'arrival_time': next_arrival_time,
                        'arrival_date': next_arrival_date,
                        'arrival_datetime': next_arrival_datetime,
                        'total_duration': new_duration,
                        'route_info': new_route_info,
                        'stops': new_stops
                    })
        
        # Filter to only routes with minimum stops
        if routes_found:
            actual_min = min(r['stops'] for r in routes_found)
            routes_found = [r for r in routes_found if r['stops'] == actual_min]
        
        return routes_found
    
    # ================================================================
    # SEARCH 2: Find FASTEST ARRIVING routes
    # Explores ALL routes and returns those arriving earliest
    # ================================================================
    def find_fastest_routes():
        """Find routes that arrive earliest at destination"""
        routes_found = []
        
        from collections import deque
        queue = deque()
        
        # Initialize with all first flights
        for first_flight in initial_flights:
            first_arrival_date = first_flight.get('arrival_date', first_flight['date'])
            first_arrival_time = first_flight['arrival']
            
            if first_arrival_time is not None:
                first_arrival_datetime = first_arrival_date.replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(minutes=first_arrival_time)
            else:
                first_arrival_datetime = first_arrival_date
            
            # Check if direct to destination
            if first_flight['destination'] == destination:
                routes_found.append({
                    'path': [origin, destination],
                    'stops': 0,
                    'total_duration': first_flight['duration'],
                    'route_info': [{
                        'from': origin,
                        'to': destination,
                        'date': first_flight['date'],
                        'departure': first_flight['dep_str'],
                        'arrival': first_flight['arr_str'],
                        'duration': first_flight['duration'],
                        'duration_str': first_flight['duration_str'],
                        'carrier': first_flight['carrier'],
                        'flight': first_flight['flight_num'],
                        'wait_time': 0
                    }],
                    'start_date': first_flight['date'],
                    'end_date': first_arrival_date,
                    'arrival_datetime': first_arrival_datetime
                })
            else:
                queue.append({
                    'current': first_flight['destination'],
                    'path': [origin, first_flight['destination']],
                    'arrival_time': first_arrival_time,
                    'arrival_date': first_arrival_date,
                    'arrival_datetime': first_arrival_datetime,
                    'total_duration': first_flight['duration'],
                    'route_info': [{
                        'from': origin,
                        'to': first_flight['destination'],
                        'date': first_flight['date'],
                        'departure': first_flight['dep_str'],
                        'arrival': first_flight['arr_str'],
                        'duration': first_flight['duration'],
                        'duration_str': first_flight['duration_str'],
                        'carrier': first_flight['carrier'],
                        'flight': first_flight['flight_num'],
                        'wait_time': 0
                    }]
                })
        
        visited = set()
        iterations = 0
        max_iterations = 100000
        
        while queue and iterations < max_iterations:
            iterations += 1
            state = queue.popleft()
            
            current = state['current']
            path = state['path']
            
            # Skip if too many stops
            if len(path) - 1 >= max_stops + 1:
                continue
            
            state_key = (current, tuple(path))
            if state_key in visited:
                continue
            visited.add(state_key)
            
            last_arrival_time = state['arrival_time']
            last_arrival_date = state['arrival_date']
            
            if current not in network:
                continue
            
            # Check ALL flights from current airport
            for next_flight in network[current]:
                if next_flight['destination'] in path:
                    continue
                
                if next_flight['date'] < last_arrival_date:
                    continue
                if next_flight['date'] > last_arrival_date + timedelta(days=3):
                    continue
                
                min_connection = 60
                
                if next_flight['date'].date() > last_arrival_date.date():
                    days_diff = (next_flight['date'].date() - last_arrival_date.date()).days
                    wait_time = (1440 - last_arrival_time) + ((days_diff - 1) * 1440) + next_flight['departure']
                elif next_flight['date'].date() == last_arrival_date.date():
                    if next_flight['departure'] >= last_arrival_time + min_connection:
                        wait_time = next_flight['departure'] - last_arrival_time
                    else:
                        continue
                else:
                    continue
                
                if wait_time < min_connection or wait_time > 1440:
                    continue
                
                next_arrival_date = next_flight.get('arrival_date', next_flight['date'])
                next_arrival_time = next_flight['arrival']
                
                if next_arrival_time is not None:
                    next_arrival_datetime = next_arrival_date.replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(minutes=next_arrival_time)
                else:
                    next_arrival_datetime = next_arrival_date
                
                new_route_info = state['route_info'] + [{
                    'from': current,
                    'to': next_flight['destination'],
                    'date': next_flight['date'],
                    'departure': next_flight['dep_str'],
                    'arrival': next_flight['arr_str'],
                    'duration': next_flight['duration'],
                    'duration_str': next_flight['duration_str'],
                    'carrier': next_flight['carrier'],
                    'flight': next_flight['flight_num'],
                    'wait_time': wait_time
                }]
                
                new_path = path + [next_flight['destination']]
                new_duration = state['total_duration'] + wait_time + next_flight['duration']
                
                if next_flight['destination'] == destination:
                    routes_found.append({
                        'path': new_path,
                        'stops': len(new_path) - 2,
                        'total_duration': new_duration,
                        'route_info': new_route_info,
                        'start_date': new_route_info[0]['date'],
                        'end_date': next_arrival_date,
                        'arrival_datetime': next_arrival_datetime
                    })
                else:
                    queue.append({
                        'current': next_flight['destination'],
                        'path': new_path,
                        'arrival_time': next_arrival_time,
                        'arrival_date': next_arrival_date,
                        'arrival_datetime': next_arrival_datetime,
                        'total_duration': new_duration,
                        'route_info': new_route_info
                    })
        
        return routes_found
    
    # Run BOTH searches
    fewest_stops_routes = find_fewest_stops_routes()
    fastest_routes = find_fastest_routes()
    
    # Combine all routes
    all_routes = fewest_stops_routes + fastest_routes
    
    # Remove duplicates
    unique_routes = []
    seen_routes = set()
    
    for route in all_routes:
        route_id = tuple([
            (leg['from'], leg['to'], leg['date'].date(), leg['departure'])
            for leg in route['route_info']
        ])
        
        if route_id not in seen_routes:
            seen_routes.add(route_id)
            unique_routes.append(route)
    
    return unique_routes

def get_fastest_and_fewest_stops_routes(all_routes_same_day):
    """
    From all routes departing on the same day:
    1. Fewest Stops: sorted by number of stops (minimum first), then by journey time
    2. Fastest Arriving: sorted by arrival_datetime (earliest first)
    
    Returns two separate lists.
    """
    if not all_routes_same_day:
        return [], []
    
    # ============================================================
    # FASTEST ARRIVING ROUTES
    # Sort by arrival time at destination (earliest first)
    # ============================================================
    fastest_arriving = sorted(all_routes_same_day, key=lambda x: (x['arrival_datetime'], x['stops']))
    
    # ============================================================
    # FEWEST STOPS ROUTES
    # Sort by number of stops (minimum first), then by total journey time
    # ============================================================
    by_fewest_stops = sorted(all_routes_same_day, key=lambda x: (x['stops'], x['total_duration']))
    
    # Find the minimum number of stops available
    min_stops = by_fewest_stops[0]['stops'] if by_fewest_stops else 0
    
    # Get all routes with minimum stops
    fewest_stops_routes = [r for r in by_fewest_stops if r['stops'] == min_stops]
    
    # Sort fewest stops routes by total journey time (shortest first)
    fewest_stops_routes = sorted(fewest_stops_routes, key=lambda x: x['total_duration'])
    
    return fastest_arriving, fewest_stops_routes

def display_route_results(origin, destination, selected_date, schedule_df, min_departure_time=0):
    """Common function to display route results
    
    Args:
        min_departure_time: Minimum departure time in minutes from midnight (e.g., 600 = 10:00 AM)
    """
    search_date = pd.Timestamp(selected_date)
    
    # Convert min_departure_time to display string
    dep_hour = min_departure_time // 60
    dep_min = min_departure_time % 60
    time_filter_str = f"{dep_hour:02d}:{dep_min:02d}" if min_departure_time > 0 else "00:00"
    
    with st.spinner(f"Searching routes from {origin} to {destination}..."):
        try:
            # Search for direct flights
            direct_results = find_direct_flights(schedule_df, origin, destination, search_date)
            
            if direct_results:
                has_same_day = any(r['days_from_requested'] == 0 for r in direct_results)
                has_earlier_arrivals = any(r.get('arrives_earlier', False) for r in direct_results)
                
                if has_same_day:
                    st.success(f"✅ Found direct flights on your selected date ({selected_date})!")
                    if has_earlier_arrivals:
                        st.info("💡 Also found flights departing later but arriving EARLIER than same-day options!")
                else:
                    st.warning(f"⚠️ No direct flights on {selected_date}. Showing next available dates.")
                
                for result in direct_results:
                    date_diff = result['days_from_requested']
                    arrives_earlier = result.get('arrives_earlier', False)
                    
                    if date_diff == 0:
                        date_label = "✓ ON YOUR SELECTED DATE"
                        color = "green"
                    elif arrives_earlier:
                        date_label = f"🌟 DEPARTS LATER BUT ARRIVES EARLIER! (+{date_diff} day(s) departure)"
                        color = "blue"
                    else:
                        date_label = f"📅 Next available: +{date_diff} day(s)"
                        color = "orange"
                    
                    st.markdown(f"""
                    <div class="route-card">
                        <h3 style="color: {color};">📅 {result['date'].strftime('%Y-%m-%d (%A)')} - {date_label}</h3>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    for i, flight in enumerate(result['flights'], 1):
                        with st.expander(f"✈️ Direct Flight Option {i} - Carrier: {flight.get('Carrier', 'N/A')} - Arrives: {flight['Sched In(L)']}", 
                                       expanded=(date_diff == 0 and i == 1)):
                            col1, col2, col3 = st.columns(3)
                            
                            with col1:
                                st.markdown("**📅 Flight Date**")
                                st.write(f"{result['date'].strftime('%Y-%m-%d')}")
                                st.write(f"{result['date'].strftime('%A')}")
                                st.markdown("**✈️ Carrier**")
                                st.write(f"{flight.get('Carrier', 'N/A')}")
                            
                            with col2:
                                st.markdown("**🕐 Schedule**")
                                st.write(f"Departure: {flight['Sched Out(L)']} from {origin}")
                                st.write(f"Arrival: {flight['Sched In(L)']} at {destination}")
                                st.markdown("**✈️ Flight Number**")
                                st.write(f"{flight.get('Carrier', '')}{flight.get('Flight #', '')}")
                            
                            with col3:
                                st.markdown("**⏱️ Duration**")
                                st.write(f"Flight Time: {flight['Blkhr']}")
                                st.write(f"No connections needed")
                            
                            st.success(f"""
                            **Summary:**
                            - Total Travel Time: {flight['Blkhr']}
                            - Direct flight (no stops)
                            - Carrier: {flight.get('Carrier', 'N/A')}
                            """)
            
            # Search for connecting flights
            if min_departure_time > 0:
                st.info(f"🔄 Searching for connecting flight options (departing from {time_filter_str} onwards)...")
            else:
                st.info("🔄 Searching for connecting flight options...")
            
            network = build_network(schedule_df, search_date)
            
            if network:
                # Find ALL routes departing on the selected date (after min_departure_time)
                all_same_day_routes = find_all_routes_for_date(network, origin, destination, search_date, min_departure_time=min_departure_time)
                
                if all_same_day_routes:
                    # Get the two sorted lists
                    fastest_routes, fewest_stops_routes = get_fastest_and_fewest_stops_routes(all_same_day_routes)
                    
                    # Cache results in session state to prevent disappearing on download
                    st.session_state.cached_routes = {
                        'fastest': fastest_routes,
                        'fewest': fewest_stops_routes,
                        'all': all_same_day_routes
                    }
                    st.session_state.cached_origin = origin
                    st.session_state.cached_destination = destination
                    st.session_state.cached_date = selected_date
                    
                    st.success(f"✅ Found {len(all_same_day_routes)} connecting route(s) departing on {selected_date}!")
                    
                    # Check for complex routes (5+ stops)
                    min_stops_overall = min(r['stops'] for r in all_same_day_routes)
                    all_routes_need_5_plus = min_stops_overall >= 5
                    
                    if all_routes_need_5_plus:
                        st.markdown("""
                        <div class="contact-warning">
                            <h3 style="color: #856404; margin-top: 0;">⚠️ Complex Routing Required</h3>
                            <p style="font-size: 16px; margin-bottom: 10px;">
                                <strong>All available routes for this origin-destination pair require 5 or more stops.</strong>
                            </p>
                            <p style="font-size: 15px; margin-bottom: 15px;">
                                For shipments requiring complex multi-stop routing, we strongly recommend contacting our logistics team 
                                for personalized assistance to ensure optimal handling and timing.
                            </p>
                            <p style="font-size: 14px; color: #666; margin-bottom: 0;">
                                📞 <strong>Please contact UPS Healthcare Logistics for assistance with this route.</strong>
                            </p>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    # ============================================================
                    # SECTION 1: ROUTES WITH FEWEST STOPS (PRIMARY)
                    # ============================================================
                    st.markdown("### 🔗 Routes with Fewest Stops & Shortest Journey")
                    st.caption(f"Routes departing on {selected_date} with minimum connections ({min_stops_overall} stop(s)), sorted by journey time")
                    
                    if min_stops_overall >= 5:
                        st.warning(f"""
                        ⚠️ **Minimum stops available: {min_stops_overall}**
                        
                        All routes for this origin-destination require {min_stops_overall} or more stops.
                        Please contact UPS Healthcare Logistics for assistance with complex routing.
                        """)
                    
                    st.info(f"""
                    💡 **Recommended**: These routes have the fewest stops ({min_stops_overall}) and shortest journey times.
                    Fewer stops = less cargo handling = reduced risk for sensitive shipments.
                    """)
                    
                    for i, route in enumerate(fewest_stops_routes[:5], 1):
                        route_str = " → ".join(route['path'])
                        total_duration = route['total_duration']
                        total_hours = total_duration // 60
                        total_mins = total_duration % 60
                        
                        total_wait = sum([leg['wait_time'] for leg in route['route_info']])
                        wait_hours = total_wait // 60
                        wait_mins = total_wait % 60
                        
                        arrival_time_str = route['arrival_datetime'].strftime('%Y-%m-%d %H:%M')
                        
                        with st.expander(f"🔗 Option {i}: {route_str} (✅ {route['stops']} stop(s), {total_hours}h {total_mins}m journey) - Arrives: {arrival_time_str}", 
                                       expanded=(i == 1)):
                            
                            if route['stops'] >= 5:
                                st.error("""
                                ⚠️ **This route requires 5 or more stops.**
                                
                                Please **contact UPS Healthcare Logistics** for personalized assistance with this complex routing.
                                """)
                            
                            # Get first flight departure time
                            first_leg = route['route_info'][0]
                            dep_time_str = first_leg['departure']
                            
                            st.markdown(f"""
                            <div style="background-color: #E8F8E8; padding: 15px; border-radius: 10px; margin-bottom: 15px; border-left: 4px solid #4CAF50;">
                                <h4 style="color: #2E7D32; margin: 0;">✅ Recommended Route - Fewest Stops & Shortest Journey</h4>
                                <p><strong>Route:</strong> {route_str}</p>
                                <p><strong>✅ Number of Stops:</strong> {route['stops']} (minimum available)</p>
                                <p><strong>🛫 Departure:</strong> {route['start_date'].strftime('%Y-%m-%d')} at {dep_time_str} ({route['start_date'].strftime('%A')})</p>
                                <p><strong>🛬 Arrival:</strong> {route['arrival_datetime'].strftime('%Y-%m-%d')} at {route['arrival_datetime'].strftime('%H:%M')} ({route['arrival_datetime'].strftime('%A')})</p>
                                <p><strong>Total Journey Time:</strong> {total_hours}h {total_mins}m</p>
                                <p><strong>Total Waiting Time:</strong> {wait_hours}h {wait_mins}m</p>
                            </div>
                            """, unsafe_allow_html=True)
                            
                            st.markdown("### ✈️ Flight Segments:")
                            
                            for j, leg in enumerate(route['route_info'], 1):
                                leg_departure_date = leg['date']
                                leg_arrival_date = leg['date']
                                
                                dep_minutes = parse_time_to_minutes(leg['departure'])
                                arr_minutes = parse_time_to_minutes(leg['arrival'])
                                if arr_minutes and dep_minutes and arr_minutes < dep_minutes:
                                    leg_arrival_date = leg['date'] + timedelta(days=1)
                                
                                st.markdown(f"""
                                <div style="background-color: #FAFAFA; padding: 15px; border-radius: 8px; margin: 10px 0; border-left: 4px solid #4CAF50;">
                                    <h4 style="color: #351C15;">Segment {j}: {leg['from']} → {leg['to']}</h4>
                                </div>
                                """, unsafe_allow_html=True)
                                
                                col1, col2, col3, col4 = st.columns(4)
                                
                                with col1:
                                    st.markdown("**Date & Carrier**")
                                    st.write(f"📅 Dep: {leg_departure_date.strftime('%Y-%m-%d')}")
                                    st.write(f"📅 Arr: {leg_arrival_date.strftime('%Y-%m-%d')}")
                                    st.write(f"✈️ Carrier: {leg['carrier']}")
                                
                                with col2:
                                    st.markdown("**Flight Details**")
                                    st.write(f"Flight: {leg['flight']}")
                                    st.write(f"Dep: {leg['departure']} ({leg_departure_date.strftime('%a')})")
                                    st.write(f"Arr: {leg['arrival']} ({leg_arrival_date.strftime('%a')})")
                                
                                with col3:
                                    st.markdown("**Duration**")
                                    st.write(f"Flight Time: {format_duration(leg['duration'])}")
                                    st.write(f"({leg['duration_str']})")
                                
                                with col4:
                                    st.markdown("**Connection**")
                                    if j < len(route['route_info']):
                                        wait_time = route['route_info'][j]['wait_time']
                                        st.write(f"⏳ Wait: {format_duration(wait_time)}")
                                    else:
                                        st.write("Final destination")
                                
                                if j < len(route['route_info']):
                                    st.markdown("⬇️")
                            
                            st.success(f"""
                            **Journey Complete (Fewest Stops):**
                            - ✅ Only {route['stops']} stop(s) - Minimum handling
                            - 🎯 Arrival Time: {arrival_time_str}
                            - Total Travel Time: {total_hours}h {total_mins}m
                            - Total Segments: {len(route['route_info'])}
                            """)
                    
                    # ============================================================
                    # SECTION 2: FASTEST ARRIVING ROUTES
                    # ============================================================
                    st.markdown("---")
                    st.markdown("### 🚀 Fastest Arriving Routes")
                    st.caption(f"Routes departing on {selected_date}, sorted by earliest arrival at {destination}")
                    
                    for i, route in enumerate(fastest_routes[:5], 1):
                        route_str = " → ".join(route['path'])
                        total_duration = route['total_duration']
                        total_hours = total_duration // 60
                        total_mins = total_duration % 60
                        
                        total_wait = sum([leg['wait_time'] for leg in route['route_info']])
                        wait_hours = total_wait // 60
                        wait_mins = total_wait % 60
                        
                        arrival_time_str = route['arrival_datetime'].strftime('%Y-%m-%d %H:%M')
                        
                        stops_display = f"{route['stops']} stop(s)"
                        if route['stops'] >= 5:
                            stops_display = f"⚠️ {route['stops']} stops - CONTACT FOR ASSISTANCE"
                        
                        with st.expander(f"🚀 Route {i}: {route_str} ({stops_display}) - Arrives: {arrival_time_str}", 
                                       expanded=False):
                            
                            if route['stops'] >= 5:
                                st.error("""
                                ⚠️ **This route requires 5 or more stops.**
                                
                                For complex multi-stop routing, please **contact UPS Healthcare Logistics** for personalized assistance.
                                """)
                            
                            # Get first flight departure time
                            first_leg = route['route_info'][0]
                            dep_time_str = first_leg['departure']
                            
                            st.markdown(f"""
                            <div style="background-color: #E8F4F8; padding: 15px; border-radius: 10px; margin-bottom: 15px;">
                                <h4 style="color: #351C15; margin: 0;">Route Summary - Fastest Arriving</h4>
                                <p><strong>Route:</strong> {route_str}</p>
                                <p><strong>🛫 Departure:</strong> {route['start_date'].strftime('%Y-%m-%d')} at {dep_time_str} ({route['start_date'].strftime('%A')})</p>
                                <p><strong>🛬 Arrival:</strong> {route['arrival_datetime'].strftime('%Y-%m-%d')} at {route['arrival_datetime'].strftime('%H:%M')} ({route['arrival_datetime'].strftime('%A')})</p>
                                <p><strong>Total Journey Time:</strong> {total_hours}h {total_mins}m</p>
                                <p><strong>Total Waiting Time:</strong> {wait_hours}h {wait_mins}m</p>
                                <p><strong>Number of Stops:</strong> {route['stops']}</p>
                            </div>
                            """, unsafe_allow_html=True)
                            
                            st.markdown("### ✈️ Flight Segments:")
                            
                            for j, leg in enumerate(route['route_info'], 1):
                                leg_departure_date = leg['date']
                                leg_arrival_date = leg['date']
                                
                                dep_minutes = parse_time_to_minutes(leg['departure'])
                                arr_minutes = parse_time_to_minutes(leg['arrival'])
                                if arr_minutes and dep_minutes and arr_minutes < dep_minutes:
                                    leg_arrival_date = leg['date'] + timedelta(days=1)
                                
                                st.markdown(f"""
                                <div style="background-color: #FAFAFA; padding: 15px; border-radius: 8px; margin: 10px 0; border-left: 4px solid #FFB500;">
                                    <h4 style="color: #351C15;">Segment {j}: {leg['from']} → {leg['to']}</h4>
                                </div>
                                """, unsafe_allow_html=True)
                                
                                col1, col2, col3, col4 = st.columns(4)
                                
                                with col1:
                                    st.markdown("**Date & Carrier**")
                                    st.write(f"📅 Dep: {leg_departure_date.strftime('%Y-%m-%d')}")
                                    st.write(f"📅 Arr: {leg_arrival_date.strftime('%Y-%m-%d')}")
                                    st.write(f"✈️ Carrier: {leg['carrier']}")
                                
                                with col2:
                                    st.markdown("**Flight Details**")
                                    st.write(f"Flight: {leg['flight']}")
                                    st.write(f"Dep: {leg['departure']} ({leg_departure_date.strftime('%a')})")
                                    st.write(f"Arr: {leg['arrival']} ({leg_arrival_date.strftime('%a')})")
                                
                                with col3:
                                    st.markdown("**Duration**")
                                    st.write(f"Flight Time: {format_duration(leg['duration'])}")
                                    st.write(f"({leg['duration_str']})")
                                
                                with col4:
                                    st.markdown("**Connection**")
                                    if j < len(route['route_info']):
                                        wait_time = route['route_info'][j]['wait_time']
                                        st.write(f"⏳ Wait: {format_duration(wait_time)}")
                                    else:
                                        st.write("Final destination")
                                
                                if j < len(route['route_info']):
                                    st.markdown("⬇️")
                            
                            st.success(f"""
                            **Journey Complete:**
                            - 🎯 Arrival Time: {arrival_time_str}
                            - Total Travel Time: {total_hours}h {total_mins}m
                            - Total Waiting Time: {wait_hours}h {wait_mins}m
                            - Total Segments: {len(route['route_info'])}
                            """)
                    
                    # ============================================================
                    # DOWNLOAD SECTION
                    # ============================================================
                    st.markdown("<br>", unsafe_allow_html=True)
                    st.markdown("""
                    <div class="download-section">
                        <h4>📥 Export Routes</h4>
                        <p style="color: #666; font-size: 0.9rem; margin-bottom: 1rem;">Download professional reports for sharing with customers and management</p>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    # Generate files and cache in session state to prevent reload
                    cache_key = f"{origin}_{destination}_{selected_date}"
                    
                    if 'pdf_cache_key' not in st.session_state or st.session_state.pdf_cache_key != cache_key:
                        # Generate and cache PDFs
                        st.session_state.pdf_cache_key = cache_key
                        st.session_state.fastest_pdf = generate_routes_pdf(
                            fastest_routes[:5],
                            origin,
                            destination,
                            selected_date,
                            route_type="fastest"
                        )
                        st.session_state.fewest_pdf = generate_routes_pdf(
                            fewest_stops_routes[:3],
                            origin,
                            destination,
                            selected_date,
                            route_type="fewest_stops"
                        )
                        # Combined report
                        fastest_paths = [tuple(r['path']) for r in fastest_routes[:5]]
                        all_routes_for_pdf = list(fastest_routes[:5])
                        for r in fewest_stops_routes[:3]:
                            if tuple(r['path']) not in fastest_paths:
                                all_routes_for_pdf.append(r)
                        st.session_state.combined_pdf = generate_routes_pdf(
                            all_routes_for_pdf,
                            origin,
                            destination,
                            selected_date,
                            route_type="fastest"
                        )
                        
                        # Generate Excel files
                        st.session_state.fastest_excel = generate_routes_excel(
                            fastest_routes[:5],
                            origin,
                            destination,
                            selected_date,
                            route_type="fastest"
                        )
                        st.session_state.fewest_excel = generate_routes_excel(
                            fewest_stops_routes[:3],
                            origin,
                            destination,
                            selected_date,
                            route_type="fewest_stops"
                        )
                        st.session_state.combined_excel = generate_routes_excel(
                            all_routes_for_pdf,
                            origin,
                            destination,
                            selected_date,
                            route_type="combined"
                        )
                    
                    # PDF Downloads
                    st.markdown("#### 📄 PDF Reports")
                    col_dl1, col_dl2 = st.columns(2)
                    
                    with col_dl1:
                        st.download_button(
                            label=f"📄 Routing {origin} to {destination} - Fewest Stops (PDF)",
                            data=st.session_state.fewest_pdf,
                            file_name=f"Routing_{origin}_to_{destination}_on_{selected_date}_Fewest_Stops.pdf",
                            mime="application/pdf",
                            use_container_width=True,
                            key=f"dl_fewest_{cache_key}"
                        )
                    
                    with col_dl2:
                        st.download_button(
                            label=f"📄 Routing {origin} to {destination} - Fastest (PDF)",
                            data=st.session_state.fastest_pdf,
                            file_name=f"Routing_{origin}_to_{destination}_on_{selected_date}_Fastest.pdf",
                            mime="application/pdf",
                            use_container_width=True,
                            key=f"dl_fastest_{cache_key}"
                        )
                    
                    st.download_button(
                        label=f"📄 Routing {origin} to {destination} - Complete Report (PDF)",
                        data=st.session_state.combined_pdf,
                        file_name=f"Routing_{origin}_to_{destination}_on_{selected_date}_Complete.pdf",
                        mime="application/pdf",
                        use_container_width=True,
                        key=f"dl_combined_{cache_key}"
                    )
                    
                    # Excel Downloads
                    st.markdown("#### 📊 Excel Reports")
                    col_xl1, col_xl2 = st.columns(2)
                    
                    with col_xl1:
                        st.download_button(
                            label=f"📊 Routing {origin} to {destination} - Fewest Stops (Excel)",
                            data=st.session_state.fewest_excel,
                            file_name=f"Routing_{origin}_to_{destination}_on_{selected_date}_Fewest_Stops.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            key=f"dl_fewest_xl_{cache_key}"
                        )
                    
                    with col_xl2:
                        st.download_button(
                            label=f"📊 Routing {origin} to {destination} - Fastest (Excel)",
                            data=st.session_state.fastest_excel,
                            file_name=f"Routing_{origin}_to_{destination}_on_{selected_date}_Fastest.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            key=f"dl_fastest_xl_{cache_key}"
                        )
                    
                    st.download_button(
                        label=f"📊 Routing {origin} to {destination} - Complete Report (Excel)",
                        data=st.session_state.combined_excel,
                        file_name=f"Routing_{origin}_to_{destination}_on_{selected_date}_Complete.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key=f"dl_combined_xl_{cache_key}"
                    )
                
                else:
                    # No routes on selected date - search for next available
                    st.warning(f"⚠️ No connecting routes departing on {selected_date}. Searching for next available dates...")
                    
                    found_alternative = False
                    for day_offset in range(1, 8):
                        alt_date = search_date + timedelta(days=day_offset)
                        alt_routes = find_all_routes_for_date(network, origin, destination, alt_date)
                        
                        if alt_routes:
                            st.success(f"✅ Found {len(alt_routes)} routes departing on **{alt_date.strftime('%Y-%m-%d (%A)')}** (+{day_offset} day(s))")
                            
                            fastest_routes, fewest_stops_routes = get_fastest_and_fewest_stops_routes(alt_routes)
                            min_stops_overall = min(r['stops'] for r in alt_routes)
                            
                            # Cache the alternative date results
                            st.session_state.cached_routes = {
                                'fastest': fastest_routes,
                                'fewest': fewest_stops_routes,
                                'all': alt_routes
                            }
                            st.session_state.cached_origin = origin
                            st.session_state.cached_destination = destination
                            st.session_state.cached_date = alt_date.strftime('%Y-%m-%d')
                            
                            # ============================================================
                            # SECTION 1: ROUTES WITH FEWEST STOPS (PRIMARY)
                            # ============================================================
                            st.markdown("### 🔗 Recommended - Fewest Stops Routes")
                            st.caption(f"Routes departing on {alt_date.strftime('%Y-%m-%d')} with minimum connections ({min_stops_overall} stop(s))")
                            
                            st.info(f"""
                            💡 **Recommended**: These routes have the fewest stops ({min_stops_overall}) and shortest journey times.
                            Fewer stops = less cargo handling = reduced risk for sensitive shipments.
                            """)
                            
                            for i, route in enumerate(fewest_stops_routes[:5], 1):
                                route_str = " → ".join(route['path'])
                                total_hours = route['total_duration'] // 60
                                total_mins = route['total_duration'] % 60
                                total_wait = sum([leg['wait_time'] for leg in route['route_info']])
                                wait_hours = total_wait // 60
                                wait_mins = total_wait % 60
                                arrival_time_str = route['arrival_datetime'].strftime('%Y-%m-%d %H:%M')
                                first_dep = route['route_info'][0]['departure']
                                
                                with st.expander(f"🔗 Option {i}: {route_str} (✅ {route['stops']} stop(s), {total_hours}h {total_mins}m journey) - Arrives: {arrival_time_str}", 
                                               expanded=(i == 1)):
                                    
                                    st.markdown(f"""
                                    <div style="background-color: #E8F8E8; padding: 15px; border-radius: 10px; margin-bottom: 15px; border-left: 4px solid #4CAF50;">
                                        <h4 style="color: #2E7D32; margin: 0;">✅ Recommended Route - Fewest Stops & Shortest Journey</h4>
                                        <p><strong>Route:</strong> {route_str}</p>
                                        <p><strong>✅ Number of Stops:</strong> {route['stops']} (minimum available)</p>
                                        <p><strong>🛫 Departure:</strong> {alt_date.strftime('%Y-%m-%d')} at {first_dep} ({alt_date.strftime('%A')})</p>
                                        <p><strong>🛬 Arrival:</strong> {route['arrival_datetime'].strftime('%Y-%m-%d')} at {route['arrival_datetime'].strftime('%H:%M')} ({route['arrival_datetime'].strftime('%A')})</p>
                                        <p><strong>Total Journey Time:</strong> {total_hours}h {total_mins}m</p>
                                        <p><strong>Total Waiting Time:</strong> {wait_hours}h {wait_mins}m</p>
                                    </div>
                                    """, unsafe_allow_html=True)
                                    
                                    st.markdown("### ✈️ Flight Segments:")
                                    for j, leg in enumerate(route['route_info'], 1):
                                        leg_departure_date = leg['date']
                                        leg_arrival_date = leg['date']
                                        dep_minutes = parse_time_to_minutes(leg['departure'])
                                        arr_minutes = parse_time_to_minutes(leg['arrival'])
                                        if arr_minutes and dep_minutes and arr_minutes < dep_minutes:
                                            leg_arrival_date = leg['date'] + timedelta(days=1)
                                        
                                        col1, col2, col3 = st.columns(3)
                                        with col1:
                                            st.write(f"**Segment {j}:** {leg['from']} → {leg['to']}")
                                            st.write(f"Flight: {leg['carrier']} {leg['flight']}")
                                        with col2:
                                            st.write(f"Dep: {leg['departure']} ({leg_departure_date.strftime('%a')})")
                                            st.write(f"Arr: {leg['arrival']} ({leg_arrival_date.strftime('%a')})")
                                        with col3:
                                            st.write(f"Duration: {leg['duration_str']}")
                                            if j < len(route['route_info']):
                                                wait_time = route['route_info'][j]['wait_time']
                                                st.write(f"Connection: {format_duration(wait_time)}")
                            
                            # ============================================================
                            # SECTION 2: FASTEST ARRIVING ROUTES
                            # ============================================================
                            st.markdown("---")
                            st.markdown("### 🚀 Fastest Arriving Routes")
                            st.caption(f"Routes departing on {alt_date.strftime('%Y-%m-%d')}, sorted by earliest arrival")
                            
                            for i, route in enumerate(fastest_routes[:5], 1):
                                route_str = " → ".join(route['path'])
                                total_hours = route['total_duration'] // 60
                                total_mins = route['total_duration'] % 60
                                arrival_time_str = route['arrival_datetime'].strftime('%Y-%m-%d %H:%M')
                                first_dep = route['route_info'][0]['departure']
                                
                                with st.expander(f"🚀 Route {i}: {route_str} ({route['stops']} stop(s)) - Arrives: {arrival_time_str}", 
                                               expanded=False):
                                    st.markdown(f"""
                                    - **Route:** {route_str}
                                    - **Stops:** {route['stops']}
                                    - **Departure:** {alt_date.strftime('%Y-%m-%d')} at {first_dep}
                                    - **Arrival:** {arrival_time_str}
                                    - **Journey Time:** {total_hours}h {total_mins}m
                                    """)
                                    
                                    for j, leg in enumerate(route['route_info'], 1):
                                        st.write(f"Segment {j}: {leg['from']} → {leg['to']} | {leg['carrier']} {leg['flight']} | Dep: {leg['departure']} Arr: {leg['arrival']}")
                            
                            # ============================================================
                            # DOWNLOAD SECTION FOR ALTERNATIVE DATE
                            # ============================================================
                            st.markdown("<br>", unsafe_allow_html=True)
                            st.markdown(f"""
                            <div class="download-section">
                                <h4>📥 Export Routes</h4>
                                <p style="color: #666; font-size: 0.9rem; margin-bottom: 1rem;">Download routes for {alt_date.strftime('%Y-%m-%d (%A)')}</p>
                            </div>
                            """, unsafe_allow_html=True)
                            
                            alt_date_str = alt_date.strftime('%Y-%m-%d')
                            cache_key = f"alt_{origin}_{destination}_{alt_date_str}"
                            
                            # Generate files for alternative date
                            if 'alt_pdf_cache_key' not in st.session_state or st.session_state.get('alt_pdf_cache_key') != cache_key:
                                st.session_state.alt_pdf_cache_key = cache_key
                                st.session_state.alt_fewest_pdf = generate_routes_pdf(
                                    fewest_stops_routes[:5], origin, destination, alt_date_str, route_type="fewest_stops"
                                )
                                st.session_state.alt_fastest_pdf = generate_routes_pdf(
                                    fastest_routes[:5], origin, destination, alt_date_str, route_type="fastest"
                                )
                                st.session_state.alt_fewest_excel = generate_routes_excel(
                                    fewest_stops_routes[:5], origin, destination, alt_date_str, route_type="fewest_stops"
                                )
                                st.session_state.alt_fastest_excel = generate_routes_excel(
                                    fastest_routes[:5], origin, destination, alt_date_str, route_type="fastest"
                                )
                            
                            # Download buttons
                            st.markdown("#### 📄 PDF Reports")
                            col_dl1, col_dl2 = st.columns(2)
                            with col_dl1:
                                st.download_button(
                                    label=f"📄 Fewest Stops - {alt_date_str} (PDF)",
                                    data=st.session_state.alt_fewest_pdf,
                                    file_name=f"Routing_{origin}_to_{destination}_on_{alt_date_str}_Fewest_Stops.pdf",
                                    mime="application/pdf",
                                    use_container_width=True,
                                    key=f"alt_dl_fewest_{cache_key}"
                                )
                            with col_dl2:
                                st.download_button(
                                    label=f"📄 Fastest Arriving - {alt_date_str} (PDF)",
                                    data=st.session_state.alt_fastest_pdf,
                                    file_name=f"Routing_{origin}_to_{destination}_on_{alt_date_str}_Fastest.pdf",
                                    mime="application/pdf",
                                    use_container_width=True,
                                    key=f"alt_dl_fastest_{cache_key}"
                                )
                            
                            st.markdown("#### 📊 Excel Reports")
                            col_xl1, col_xl2 = st.columns(2)
                            with col_xl1:
                                st.download_button(
                                    label=f"📊 Fewest Stops - {alt_date_str} (Excel)",
                                    data=st.session_state.alt_fewest_excel,
                                    file_name=f"Routing_{origin}_to_{destination}_on_{alt_date_str}_Fewest_Stops.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True,
                                    key=f"alt_dl_fewest_xl_{cache_key}"
                                )
                            with col_xl2:
                                st.download_button(
                                    label=f"📊 Fastest Arriving - {alt_date_str} (Excel)",
                                    data=st.session_state.alt_fastest_excel,
                                    file_name=f"Routing_{origin}_to_{destination}_on_{alt_date_str}_Fastest.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    use_container_width=True,
                                    key=f"alt_dl_fastest_xl_{cache_key}"
                                )
                            
                            found_alternative = True
                            break
                    
                    if not found_alternative and not direct_results:
                        st.error(f"""
                        ❌ No routes found from {origin} to {destination}
                        
                        **Suggestions:**
                        - This route may not be served by UPS flights
                        - Try selecting a different origin-destination pair
                        - Check if flights operate on different days of the week
                        """)
            else:
                if not direct_results:
                    st.error("No flight network available for the selected date range.")
        
        except Exception as e:
            st.error(f"Error during search: {str(e)}")
            import traceback
            st.error(traceback.format_exc())

def display_radiopharma_results(origin, destination, selected_date, schedule_df, rp_config, min_departure_time=0):
    """Display RadioPharma route results with specific constraints"""
    search_date = pd.Timestamp(selected_date)
    
    # Convert min_departure_time to display string
    dep_hour = min_departure_time // 60
    dep_min = min_departure_time % 60
    time_filter_str = f"{dep_hour:02d}:{dep_min:02d}" if min_departure_time > 0 else "00:00"
    
    with st.spinner(f"Searching RadioPharma routes from {origin} to {destination}..."):
        try:
            # Build RadioPharma-specific network (excludes prohibited flights)
            if min_departure_time > 0:
                st.info(f"🔄 Searching RadioPharma routes (departing from {time_filter_str} onwards)...")
            else:
                st.info("🔄 Searching RadioPharma routes (excluding prohibited flights)...")
            
            network = build_radiopharma_network(schedule_df, search_date, rp_config)
            
            if network:
                # Find routes using RadioPharma-specific function
                all_same_day_routes = find_radiopharma_routes_for_date(
                    network, origin, destination, search_date, rp_config,
                    min_departure_time=min_departure_time
                )
                
                if all_same_day_routes:
                    # Get the two sorted lists
                    fastest_routes, fewest_stops_routes = get_fastest_and_fewest_stops_routes(all_same_day_routes)
                    
                    # Cache results in session state
                    st.session_state.cached_routes = {
                        'fastest': fastest_routes,
                        'fewest': fewest_stops_routes,
                        'all': all_same_day_routes
                    }
                    st.session_state.cached_origin = origin
                    st.session_state.cached_destination = destination
                    st.session_state.cached_date = selected_date
                    
                    st.success(f"✅ Found {len(all_same_day_routes)} RadioPharma route(s) departing on {selected_date}!")
                    
                    # Check for complex routes (5+ stops)
                    min_stops_overall = min(r['stops'] for r in all_same_day_routes)
                    all_routes_need_5_plus = min_stops_overall >= 5
                    
                    if all_routes_need_5_plus:
                        st.markdown("""
                        <div class="contact-warning">
                            <h3 style="color: #856404; margin-top: 0;">⚠️ Complex Routing Required</h3>
                            <p style="font-size: 16px; margin-bottom: 10px;">
                                <strong>All available RadioPharma routes require 5 or more stops.</strong>
                            </p>
                            <p style="font-size: 14px; color: #666; margin-bottom: 0;">
                                📞 <strong>Please contact UPS Healthcare Logistics for assistance.</strong>
                            </p>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    # ============================================================
                    # SECTION 1: ROUTES WITH FEWEST STOPS (PRIMARY)
                    # ============================================================
                    st.markdown("### 🔗 RadioPharma Routes with Fewest Stops & Shortest Journey")
                    st.caption(f"Routes departing on {selected_date} with minimum connections ({min_stops_overall} stop(s)), sorted by journey time")
                    
                    if min_stops_overall >= 5:
                        st.warning(f"""
                        ⚠️ **Minimum stops available: {min_stops_overall}**
                        
                        All RadioPharma routes require {min_stops_overall} or more stops.
                        Please contact UPS Healthcare Logistics for assistance.
                        """)
                    
                    st.info(f"""
                    💡 **Recommended**: These routes have the fewest stops ({min_stops_overall}) and shortest journey times.
                    Fewer stops = less cargo handling = reduced risk for sensitive radiopharmaceutical shipments.
                    """)
                    
                    for i, route in enumerate(fewest_stops_routes[:5], 1):
                        route_str = " → ".join(route['path'])
                        total_duration = route['total_duration']
                        total_hours = total_duration // 60
                        total_mins = total_duration % 60
                        
                        total_wait = sum([leg['wait_time'] for leg in route['route_info']])
                        wait_hours = total_wait // 60
                        wait_mins = total_wait % 60
                        
                        arrival_time_str = route['arrival_datetime'].strftime('%Y-%m-%d %H:%M')
                        
                        with st.expander(f"🔗 Option {i}: {route_str} (✅ {route['stops']} stop(s), {total_hours}h {total_mins}m journey) - Arrives: {arrival_time_str}", 
                                       expanded=(i == 1)):
                            
                            if route['stops'] >= 5:
                                st.error("""
                                ⚠️ **This route requires 5 or more stops.**
                                
                                Please **contact UPS Healthcare Logistics** for assistance.
                                """)
                            
                            first_leg = route['route_info'][0]
                            dep_time_str = first_leg['departure']
                            
                            st.markdown(f"""
                            <div style="background-color: #E8F8E8; padding: 15px; border-radius: 10px; margin-bottom: 15px; border-left: 4px solid #4CAF50;">
                                <h4 style="color: #2E7D32; margin: 0;">☢️ Recommended RadioPharma Route - Fewest Stops & Shortest Journey</h4>
                                <p><strong>Route:</strong> {route_str}</p>
                                <p><strong>✅ Number of Stops:</strong> {route['stops']} (minimum available)</p>
                                <p><strong>🛫 Departure:</strong> {route['start_date'].strftime('%Y-%m-%d')} at {dep_time_str} ({route['start_date'].strftime('%A')})</p>
                                <p><strong>🛬 Arrival:</strong> {route['arrival_datetime'].strftime('%Y-%m-%d')} at {route['arrival_datetime'].strftime('%H:%M')} ({route['arrival_datetime'].strftime('%A')})</p>
                                <p><strong>Total Journey Time:</strong> {total_hours}h {total_mins}m</p>
                                <p><strong>Total Waiting Time:</strong> {wait_hours}h {wait_mins}m</p>
                            </div>
                            """, unsafe_allow_html=True)
                            
                            st.markdown("### ✈️ Flight Segments:")
                            
                            for j, leg in enumerate(route['route_info'], 1):
                                leg_departure_date = leg['date']
                                leg_arrival_date = leg['date']
                                
                                dep_minutes = parse_time_to_minutes(leg['departure'])
                                arr_minutes = parse_time_to_minutes(leg['arrival'])
                                if arr_minutes and dep_minutes and arr_minutes < dep_minutes:
                                    leg_arrival_date = leg['date'] + timedelta(days=1)
                                
                                st.markdown(f"""
                                <div style="background-color: #FAFAFA; padding: 15px; border-radius: 8px; margin: 10px 0; border-left: 4px solid #4CAF50;">
                                    <h4 style="color: #351C15;">Segment {j}: {leg['from']} → {leg['to']}</h4>
                                </div>
                                """, unsafe_allow_html=True)
                                
                                col1, col2, col3, col4 = st.columns(4)
                                
                                with col1:
                                    st.markdown("**Date & Carrier**")
                                    st.write(f"📅 Dep: {leg_departure_date.strftime('%Y-%m-%d')}")
                                    st.write(f"📅 Arr: {leg_arrival_date.strftime('%Y-%m-%d')}")
                                    st.write(f"✈️ Carrier: {leg['carrier']}")
                                
                                with col2:
                                    st.markdown("**Flight Details**")
                                    st.write(f"Flight: {leg['flight']}")
                                    st.write(f"Dep: {leg['departure']} ({leg_departure_date.strftime('%a')})")
                                    st.write(f"Arr: {leg['arrival']} ({leg_arrival_date.strftime('%a')})")
                                
                                with col3:
                                    st.markdown("**Duration**")
                                    st.write(f"Flight Time: {format_duration(leg['duration'])}")
                                    st.write(f"({leg['duration_str']})")
                                
                                with col4:
                                    st.markdown("**Connection**")
                                    if j < len(route['route_info']):
                                        wait_time = route['route_info'][j]['wait_time']
                                        st.write(f"⏳ Wait: {format_duration(wait_time)}")
                                    else:
                                        st.write("Final destination")
                                
                                if j < len(route['route_info']):
                                    st.markdown("⬇️")
                            
                            st.success(f"""
                            **Journey Complete (Fewest Stops):**
                            - ✅ Only {route['stops']} stop(s) - Minimum handling
                            - 🎯 Arrival Time: {arrival_time_str}
                            - Total Travel Time: {total_hours}h {total_mins}m
                            - Total Segments: {len(route['route_info'])}
                            """)
                    
                    # ============================================================
                    # SECTION 2: FASTEST ARRIVING ROUTES
                    # ============================================================
                    st.markdown("---")
                    st.markdown("### 🚀 Fastest Arriving RadioPharma Routes")
                    st.caption(f"Routes departing on {selected_date}, sorted by earliest arrival at {destination}")
                    
                    for i, route in enumerate(fastest_routes[:5], 1):
                        route_str = " → ".join(route['path'])
                        total_duration = route['total_duration']
                        total_hours = total_duration // 60
                        total_mins = total_duration % 60
                        
                        total_wait = sum([leg['wait_time'] for leg in route['route_info']])
                        wait_hours = total_wait // 60
                        wait_mins = total_wait % 60
                        
                        arrival_time_str = route['arrival_datetime'].strftime('%Y-%m-%d %H:%M')
                        
                        stops_display = f"{route['stops']} stop(s)"
                        if route['stops'] >= 5:
                            stops_display = f"⚠️ {route['stops']} stops - CONTACT FOR ASSISTANCE"
                        
                        with st.expander(f"🚀 Route {i}: {route_str} ({stops_display}) - Arrives: {arrival_time_str}", 
                                       expanded=False):
                            
                            if route['stops'] >= 5:
                                st.error("""
                                ⚠️ **This route requires 5 or more stops.**
                                
                                Please **contact UPS Healthcare Logistics** for personalized assistance.
                                """)
                            
                            first_leg = route['route_info'][0]
                            dep_time_str = first_leg['departure']
                            
                            st.markdown(f"""
                            <div style="background-color: #E8F4F8; padding: 15px; border-radius: 10px; margin-bottom: 15px;">
                                <h4 style="color: #351C15; margin: 0;">☢️ RadioPharma Route Summary - Fastest Arriving</h4>
                                <p><strong>Route:</strong> {route_str}</p>
                                <p><strong>🛫 Departure:</strong> {route['start_date'].strftime('%Y-%m-%d')} at {dep_time_str} ({route['start_date'].strftime('%A')})</p>
                                <p><strong>🛬 Arrival:</strong> {route['arrival_datetime'].strftime('%Y-%m-%d')} at {route['arrival_datetime'].strftime('%H:%M')} ({route['arrival_datetime'].strftime('%A')})</p>
                                <p><strong>Total Journey Time:</strong> {total_hours}h {total_mins}m</p>
                                <p><strong>Total Waiting Time:</strong> {wait_hours}h {wait_mins}m</p>
                                <p><strong>Number of Stops:</strong> {route['stops']}</p>
                            </div>
                            """, unsafe_allow_html=True)
                            
                            st.markdown("### ✈️ Flight Segments:")
                            
                            for j, leg in enumerate(route['route_info'], 1):
                                leg_departure_date = leg['date']
                                leg_arrival_date = leg['date']
                                
                                dep_minutes = parse_time_to_minutes(leg['departure'])
                                arr_minutes = parse_time_to_minutes(leg['arrival'])
                                if arr_minutes and dep_minutes and arr_minutes < dep_minutes:
                                    leg_arrival_date = leg['date'] + timedelta(days=1)
                                
                                st.markdown(f"""
                                <div style="background-color: #FAFAFA; padding: 15px; border-radius: 8px; margin: 10px 0; border-left: 4px solid #FFB500;">
                                    <h4 style="color: #351C15;">Segment {j}: {leg['from']} → {leg['to']}</h4>
                                </div>
                                """, unsafe_allow_html=True)
                                
                                col1, col2, col3, col4 = st.columns(4)
                                
                                with col1:
                                    st.markdown("**Date & Carrier**")
                                    st.write(f"📅 Dep: {leg_departure_date.strftime('%Y-%m-%d')}")
                                    st.write(f"📅 Arr: {leg_arrival_date.strftime('%Y-%m-%d')}")
                                    st.write(f"✈️ Carrier: {leg['carrier']}")
                                
                                with col2:
                                    st.markdown("**Flight Details**")
                                    st.write(f"Flight: {leg['flight']}")
                                    st.write(f"Dep: {leg['departure']} ({leg_departure_date.strftime('%a')})")
                                    st.write(f"Arr: {leg['arrival']} ({leg_arrival_date.strftime('%a')})")
                                
                                with col3:
                                    st.markdown("**Duration**")
                                    st.write(f"Flight Time: {format_duration(leg['duration'])}")
                                    st.write(f"({leg['duration_str']})")
                                
                                with col4:
                                    st.markdown("**Connection**")
                                    if j < len(route['route_info']):
                                        wait_time = route['route_info'][j]['wait_time']
                                        st.write(f"⏳ Wait: {format_duration(wait_time)}")
                                    else:
                                        st.write("Final destination")
                                
                                if j < len(route['route_info']):
                                    st.markdown("⬇️")
                            
                            st.success(f"""
                            **Journey Complete:**
                            - 🎯 Arrival Time: {arrival_time_str}
                            - Total Travel Time: {total_hours}h {total_mins}m
                            - Total Waiting Time: {wait_hours}h {wait_mins}m
                            - Total Segments: {len(route['route_info'])}
                            """)
                    
                    # ============================================================
                    # DOWNLOAD SECTION
                    # ============================================================
                    st.markdown("<br>", unsafe_allow_html=True)
                    st.markdown("""
                    <div class="download-section">
                        <h4>📥 Export RadioPharma Routes</h4>
                        <p style="color: #666; font-size: 0.9rem; margin-bottom: 1rem;">Download compliance-checked reports for sharing with customers and management</p>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    cache_key = f"rp_{origin}_{destination}_{selected_date}"
                    
                    if 'rp_pdf_cache_key' not in st.session_state or st.session_state.rp_pdf_cache_key != cache_key:
                        st.session_state.rp_pdf_cache_key = cache_key
                        st.session_state.rp_fastest_pdf = generate_routes_pdf(
                            fastest_routes[:5],
                            origin,
                            destination,
                            selected_date,
                            route_type="fastest"
                        )
                        st.session_state.rp_fewest_pdf = generate_routes_pdf(
                            fewest_stops_routes[:3],
                            origin,
                            destination,
                            selected_date,
                            route_type="fewest_stops"
                        )
                        fastest_paths = [tuple(r['path']) for r in fastest_routes[:5]]
                        all_routes_for_pdf = list(fastest_routes[:5])
                        for r in fewest_stops_routes[:3]:
                            if tuple(r['path']) not in fastest_paths:
                                all_routes_for_pdf.append(r)
                        st.session_state.rp_combined_pdf = generate_routes_pdf(
                            all_routes_for_pdf,
                            origin,
                            destination,
                            selected_date,
                            route_type="fastest"
                        )
                        
                        # Generate Excel files
                        st.session_state.rp_fastest_excel = generate_routes_excel(
                            fastest_routes[:5],
                            origin,
                            destination,
                            selected_date,
                            route_type="fastest"
                        )
                        st.session_state.rp_fewest_excel = generate_routes_excel(
                            fewest_stops_routes[:3],
                            origin,
                            destination,
                            selected_date,
                            route_type="fewest_stops"
                        )
                        st.session_state.rp_combined_excel = generate_routes_excel(
                            all_routes_for_pdf,
                            origin,
                            destination,
                            selected_date,
                            route_type="combined"
                        )
                    
                    # PDF Downloads
                    st.markdown("#### 📄 PDF Reports")
                    col_dl1, col_dl2 = st.columns(2)
                    
                    with col_dl1:
                        st.download_button(
                            label=f"📄 Routing {origin} to {destination} - Fewest Stops (PDF)",
                            data=st.session_state.rp_fewest_pdf,
                            file_name=f"RadioPharma_Routing_{origin}_to_{destination}_on_{selected_date}_Fewest_Stops.pdf",
                            mime="application/pdf",
                            use_container_width=True,
                            key=f"rp_dl_fewest_{cache_key}"
                        )
                    
                    with col_dl2:
                        st.download_button(
                            label=f"📄 Routing {origin} to {destination} - Fastest (PDF)",
                            data=st.session_state.rp_fastest_pdf,
                            file_name=f"RadioPharma_Routing_{origin}_to_{destination}_on_{selected_date}_Fastest.pdf",
                            mime="application/pdf",
                            use_container_width=True,
                            key=f"rp_dl_fastest_{cache_key}"
                        )
                    
                    st.download_button(
                        label=f"📄 Routing {origin} to {destination} - Complete Report (PDF)",
                        data=st.session_state.rp_combined_pdf,
                        file_name=f"RadioPharma_Routing_{origin}_to_{destination}_on_{selected_date}_Complete.pdf",
                        mime="application/pdf",
                        use_container_width=True,
                        key=f"rp_dl_combined_{cache_key}"
                    )
                    
                    # Excel Downloads
                    st.markdown("#### 📊 Excel Reports")
                    col_xl1, col_xl2 = st.columns(2)
                    
                    with col_xl1:
                        st.download_button(
                            label=f"📊 Routing {origin} to {destination} - Fewest Stops (Excel)",
                            data=st.session_state.rp_fewest_excel,
                            file_name=f"RadioPharma_Routing_{origin}_to_{destination}_on_{selected_date}_Fewest_Stops.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            key=f"rp_dl_fewest_xl_{cache_key}"
                        )
                    
                    with col_xl2:
                        st.download_button(
                            label=f"📊 Routing {origin} to {destination} - Fastest (Excel)",
                            data=st.session_state.rp_fastest_excel,
                            file_name=f"RadioPharma_Routing_{origin}_to_{destination}_on_{selected_date}_Fastest.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            key=f"rp_dl_fastest_xl_{cache_key}"
                        )
                    
                    st.download_button(
                        label=f"📊 Routing {origin} to {destination} - Complete Report (Excel)",
                        data=st.session_state.rp_combined_excel,
                        file_name=f"RadioPharma_Routing_{origin}_to_{destination}_on_{selected_date}_Complete.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key=f"rp_dl_combined_xl_{cache_key}"
                    )
                
                else:
                    # No routes on selected date - search for next available
                    st.warning(f"⚠️ No RadioPharma routes found departing on {selected_date}. Searching for next available dates...")
                    
                    found_alternative = False
                    for day_offset in range(1, 8):
                        alt_date = search_date + timedelta(days=day_offset)
                        alt_network = build_radiopharma_network(schedule_df, alt_date, rp_config)
                        
                        if alt_network:
                            alt_routes = find_radiopharma_routes_for_date(
                                alt_network, origin, destination, alt_date, rp_config,
                                min_departure_time=min_departure_time
                            )
                            
                            if alt_routes:
                                st.success(f"✅ Found {len(alt_routes)} RadioPharma routes departing on **{alt_date.strftime('%Y-%m-%d (%A)')}** (+{day_offset} day(s))")
                                
                                fastest_routes, fewest_stops_routes = get_fastest_and_fewest_stops_routes(alt_routes)
                                min_stops_overall = min(r['stops'] for r in alt_routes)
                                
                                # Show fewest stops routes
                                st.markdown("### 🔗 Recommended - Fewest Stops RadioPharma Routes")
                                st.caption(f"Routes departing on {alt_date.strftime('%Y-%m-%d')} with minimum connections")
                                
                                for i, route in enumerate(fewest_stops_routes[:3], 1):
                                    route_str = " → ".join(route['path'])
                                    total_hours = route['total_duration'] // 60
                                    total_mins = route['total_duration'] % 60
                                    arrival_time_str = route['arrival_datetime'].strftime('%Y-%m-%d %H:%M')
                                    first_dep = route['route_info'][0]['departure']
                                    
                                    with st.expander(f"☢️ Option {i}: {route_str} (✅ {route['stops']} stop(s)) - Arrives: {arrival_time_str}", 
                                                   expanded=(i == 1)):
                                        st.markdown(f"""
                                        - **Route:** {route_str}
                                        - **Stops:** {route['stops']} (minimum)
                                        - **Departure:** {alt_date.strftime('%Y-%m-%d')} at {first_dep}
                                        - **Arrival:** {arrival_time_str}
                                        - **Journey Time:** {total_hours}h {total_mins}m
                                        """)
                                        
                                        for j, leg in enumerate(route['route_info'], 1):
                                            st.write(f"Segment {j}: {leg['from']} → {leg['to']} | {leg['carrier']} {leg['flight']} | Dep: {leg['departure']} Arr: {leg['arrival']}")
                                
                                # Download section
                                st.markdown("---")
                                st.markdown("### 📥 Export RadioPharma Routes")
                                
                                alt_date_str = alt_date.strftime('%Y-%m-%d')
                                rp_alt_cache_key = f"rp_alt_{origin}_{destination}_{alt_date_str}"
                                
                                if 'rp_alt_pdf_cache_key' not in st.session_state or st.session_state.get('rp_alt_pdf_cache_key') != rp_alt_cache_key:
                                    st.session_state.rp_alt_pdf_cache_key = rp_alt_cache_key
                                    st.session_state.rp_alt_fewest_pdf = generate_routes_pdf(
                                        fewest_stops_routes[:5], origin, destination, alt_date_str, route_type="fewest_stops"
                                    )
                                    st.session_state.rp_alt_fewest_excel = generate_routes_excel(
                                        fewest_stops_routes[:5], origin, destination, alt_date_str, route_type="fewest_stops"
                                    )
                                
                                col_dl1, col_dl2 = st.columns(2)
                                with col_dl1:
                                    st.download_button(
                                        label=f"📄 RadioPharma Routes - {alt_date_str} (PDF)",
                                        data=st.session_state.rp_alt_fewest_pdf,
                                        file_name=f"RadioPharma_Routing_{origin}_to_{destination}_on_{alt_date_str}.pdf",
                                        mime="application/pdf",
                                        use_container_width=True,
                                        key=f"rp_alt_dl_pdf_{rp_alt_cache_key}"
                                    )
                                with col_dl2:
                                    st.download_button(
                                        label=f"📊 RadioPharma Routes - {alt_date_str} (Excel)",
                                        data=st.session_state.rp_alt_fewest_excel,
                                        file_name=f"RadioPharma_Routing_{origin}_to_{destination}_on_{alt_date_str}.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        use_container_width=True,
                                        key=f"rp_alt_dl_xl_{rp_alt_cache_key}"
                                    )
                                
                                found_alternative = True
                                break
                    
                    if not found_alternative:
                        st.error(f"""
                        ❌ No RadioPharma routes found from {origin} to {destination} in the next 7 days.
                        
                        **Possible reasons:**
                        - No approved flights available
                        - All possible routes use prohibited flights
                        - No valid transit points available
                        
                        Please contact UPS Healthcare Logistics for assistance.
                        """)
            else:
                st.error("No RadioPharma flight network available for the selected date range.")
        
        except Exception as e:
            st.error(f"Error during RadioPharma search: {str(e)}")
            import traceback
            st.error(traceback.format_exc())

# Main Application
def main():
    # Sidebar - Professional Enterprise Design
    with st.sidebar:
        # Brand header
        st.markdown("""
        <div style="background: rgba(255, 181, 0, 0.08); border: 1px solid rgba(255, 181, 0, 0.15); 
                    border-radius: 8px; padding: 16px; margin-bottom: 24px;">
            <div style="display: flex; align-items: center; gap: 12px;">
                <div style="width: 36px; height: 36px; background: #FFB500; border-radius: 6px; 
                            display: flex; align-items: center; justify-content: center;">
                    <span style="font-size: 18px;">✈</span>
                </div>
                <div>
                    <div style="color: #FFB500; font-size: 14px; font-weight: 600;">Flight Routing</div>
                    <div style="color: rgba(255,255,255,0.5); font-size: 11px;">UPS Healthcare</div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        # Data Upload Section
        st.markdown("""
        <div style="color: rgba(255,255,255,0.5); font-size: 10px; font-weight: 600; 
                    text-transform: uppercase; letter-spacing: 1px; margin-bottom: 8px;">
            Data Source
        </div>
        """, unsafe_allow_html=True)
        
        uploaded_file = st.file_uploader(
            "Upload Flight Schedule",
            type=['xlsx', 'xls'],
            help="Upload your UPS Flight Schedule Excel file",
            label_visibility="collapsed"
        )
        
        if uploaded_file:
            st.markdown("""
            <div style="background: rgba(0, 135, 90, 0.12); border: 1px solid rgba(0, 135, 90, 0.25); 
                        color: #57d9a3; padding: 10px 14px; border-radius: 6px; margin-top: 12px;
                        display: flex; align-items: center; gap: 8px; font-size: 13px;">
                <span style="font-weight: 600;">✓</span>
                <span>File loaded</span>
            </div>
            """, unsafe_allow_html=True)
            
            st.markdown("<div style='height: 20px;'></div>", unsafe_allow_html=True)
            
            # File Info Section
            st.markdown("""
            <div style="color: rgba(255,255,255,0.5); font-size: 10px; font-weight: 600; 
                        text-transform: uppercase; letter-spacing: 1px; margin-bottom: 8px;">
                File Structure
            </div>
            <div style="background: rgba(255,255,255,0.03); padding: 12px; border-radius: 6px; 
                        border: 1px solid rgba(255,255,255,0.06); font-size: 12px;">
                <div style="color: rgba(255,255,255,0.7); line-height: 1.7;">
                    <div style="display: flex; align-items: center; gap: 6px;">
                        <span style="color: #FFB500; font-size: 8px;">●</span> Flight Schedules
                    </div>
                    <div style="display: flex; align-items: center; gap: 6px;">
                        <span style="color: #FFB500; font-size: 8px;">●</span> Route Pairs (Data)
                    </div>
                    <div style="display: flex; align-items: center; gap: 6px; color: rgba(255,255,255,0.4);">
                        <span style="font-size: 8px;">○</span> RP Info (optional)
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        else:
            st.markdown("""
            <div style="background: rgba(255,181,0,0.05); padding: 20px 16px; border-radius: 8px; 
                        text-align: center; margin-top: 12px; border: 1px dashed rgba(255,181,0,0.2);">
                <div style="font-size: 24px; margin-bottom: 8px; opacity: 0.6;">↑</div>
                <div style="color: rgba(255,255,255,0.6); font-size: 12px; line-height: 1.5;">
                    Upload Excel file to begin<br>
                    <span style="color: rgba(255,255,255,0.35); font-size: 11px;">Drag & drop or click</span>
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        # Footer
        st.markdown("""
        <div style="position: fixed; bottom: 0; left: 0; width: inherit; padding: 12px 16px; 
                    text-align: center; background: linear-gradient(180deg, transparent 0%, #1a0e0a 40%);
                    border-top: 1px solid rgba(255,255,255,0.03);">
            <div style="color: rgba(255,255,255,0.25); font-size: 10px; letter-spacing: 0.02em;">
                UPS Healthcare Logistics<br>
                <span style="color: rgba(255,181,0,0.4);">© 2025 Marken</span>
            </div>
        </div>
        """, unsafe_allow_html=True)
    
    # Main content
    if uploaded_file:
        with st.spinner("Loading flight data..."):
            schedule_df, routes_df, rp_config = load_data(uploaded_file)
        
        if schedule_df is not None and routes_df is not None:
            # Statistics Cards - Enterprise KPIs
            st.markdown("""
            <div style="display: flex; align-items: center; gap: 8px; margin-bottom: 16px;">
                <div style="font-size: 16px; font-weight: 600; color: #172b4d;">Network Overview</div>
                <div style="flex: 1; height: 1px; background: #e8eaed;"></div>
            </div>
            """, unsafe_allow_html=True)
            
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.markdown(f"""
                <div class="kpi-card">
                    <div class="kpi-value">{len(schedule_df):,}</div>
                    <div class="kpi-label">Total Flights</div>
                </div>
                """, unsafe_allow_html=True)
            
            with col2:
                st.markdown(f"""
                <div class="kpi-card">
                    <div class="kpi-value">{len(routes_df):,}</div>
                    <div class="kpi-label">Route Pairs</div>
                </div>
                """, unsafe_allow_html=True)
            
            with col3:
                st.markdown(f"""
                <div class="kpi-card">
                    <div class="kpi-value">{schedule_df['Orig'].nunique()}</div>
                    <div class="kpi-label">Airports</div>
                </div>
                """, unsafe_allow_html=True)
            
            with col4:
                carriers = schedule_df['Carrier'].nunique() if 'Carrier' in schedule_df.columns else 0
                st.markdown(f"""
                <div class="kpi-card">
                    <div class="kpi-value">{carriers}</div>
                    <div class="kpi-label">Carriers</div>
                </div>
                """, unsafe_allow_html=True)
            
            st.markdown("<div style='height: 24px;'></div>", unsafe_allow_html=True)
            
            # Tab Navigation - Enterprise Style
            st.markdown("""
            <div style="display: flex; align-items: center; gap: 8px; margin-bottom: 16px;">
                <div style="font-size: 16px; font-weight: 600; color: #172b4d;">Route Finder</div>
                <div style="flex: 1; height: 1px; background: #e8eaed;"></div>
            </div>
            """, unsafe_allow_html=True)
            
            tab_options = ["Tracked Routes", "Custom Routes", "RadioPharma"]
            
            # Get current index
            if st.session_state.selected_tab in tab_options:
                current_index = tab_options.index(st.session_state.selected_tab)
            else:
                current_index = 0
            
            selected_tab = st.radio(
                "Select route type:",
                tab_options,
                horizontal=True,
                key="tab_radio",
                index=current_index,
                label_visibility="collapsed"
            )
            st.session_state.selected_tab = selected_tab
            
            st.markdown("<br>", unsafe_allow_html=True)
            
            # Tab Content Container
            st.markdown('<div class="section-card">', unsafe_allow_html=True)
            
            # Tab 1: Tracked Routes
            if selected_tab == "Tracked Routes":
                st.markdown("""
                <div style="margin-bottom: 20px;">
                    <div style="font-size: 15px; font-weight: 600; color: #172b4d; margin-bottom: 4px;">Tracked Route Finder</div>
                    <div style="font-size: 13px; color: #6b778c;">Select from pre-defined route pairs configured in your data sheet</div>
                </div>
                """, unsafe_allow_html=True)
                
                col1, col2 = st.columns(2)
                
                with col1:
                    route_pairs = routes_df[['Origin Airport', 'Destination Airport']].drop_duplicates()
                    route_pairs = route_pairs.dropna()
                    
                    route_options = []
                    route_dict = {}
                    for idx, row in route_pairs.iterrows():
                        route_str = f"{row['Origin Airport']} → {row['Destination Airport']}"
                        route_options.append(route_str)
                        route_dict[route_str] = (row['Origin Airport'], row['Destination Airport'])
                    
                    selected_route = st.selectbox(
                        "Select Origin → Destination Route",
                        options=sorted(route_options),
                        help="Select from available route pairs",
                        key="tracked_route"
                    )
                    
                    if selected_route:
                        origin, destination = route_dict[selected_route]
                        st.markdown(f"""
                        <div style="background-color: #FFF8E8; padding: 10px; border-radius: 5px; border-left: 3px solid #FFB500;">
                            <strong>Selected Route:</strong> {origin} → {destination}
                        </div>
                        """, unsafe_allow_html=True)
                
                with col2:
                    min_date = schedule_df['Start Date (LZ)'].min()
                    max_date = schedule_df['End Date (LZ)'].max()
                    
                    if pd.notna(min_date) and pd.notna(max_date):
                        selected_date = st.date_input(
                            "Select Shipment Date",
                            value=min_date.date(),
                            min_value=min_date.date(),
                            max_value=max_date.date(),
                            key="tracked_date"
                        )
                        
                        day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
                        day_of_week = day_names[selected_date.weekday()]
                        st.markdown(f"""
                        <div style="background-color: #FFF8E8; padding: 10px; border-radius: 5px; border-left: 3px solid #FFB500;">
                            <strong>Selected Date:</strong> {selected_date} ({day_of_week})
                        </div>
                        """, unsafe_allow_html=True)
                
                # Time filter option
                st.markdown("---")
                use_time_filter = st.checkbox(
                    "⏰ Set earliest departure time",
                    help="Only show flights departing after a specific time",
                    key="tracked_use_time_filter"
                )
                
                tracked_min_departure = 0
                if use_time_filter:
                    st.caption("Show only flights departing from:")
                    col_t1, col_t2 = st.columns(2)
                    with col_t1:
                        tracked_dep_hour = st.selectbox(
                            "Hour",
                            options=list(range(24)),
                            format_func=lambda x: f"{x:02d}:00",
                            key="tracked_dep_hour"
                        )
                    with col_t2:
                        tracked_dep_min = st.selectbox(
                            "Minutes",
                            options=[0, 15, 30, 45],
                            format_func=lambda x: f":{x:02d}",
                            key="tracked_dep_min"
                        )
                    tracked_min_departure = tracked_dep_hour * 60 + tracked_dep_min
                    st.info(f"🕐 Showing flights departing **{tracked_dep_hour:02d}:{tracked_dep_min:02d}** or later")
                
                st.markdown("<br>", unsafe_allow_html=True)
                
                if st.button("🔍 Find Available Routes", type="primary", use_container_width=True, key="tracked_search"):
                    if selected_route:
                        display_route_results(origin, destination, selected_date, schedule_df, min_departure_time=tracked_min_departure)
            
            # Tab 2: Custom Routes
            elif selected_tab == "Custom Routes":
                st.markdown("""
                <div style="margin-bottom: 20px;">
                    <div style="font-size: 15px; font-weight: 600; color: #172b4d; margin-bottom: 4px;">Custom Route Finder</div>
                    <div style="font-size: 13px; color: #6b778c;">Select any origin and destination from all available airports</div>
                </div>
                """, unsafe_allow_html=True)
                
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    origins = sorted(schedule_df['Orig'].dropna().unique())
                    custom_origin = st.selectbox(
                        "Select Origin Airport",
                        options=origins,
                        help="Select any available origin airport",
                        key="custom_origin"
                    )
                
                with col2:
                    destinations = sorted(schedule_df['Dest'].dropna().unique())
                    custom_destination = st.selectbox(
                        "Select Destination Airport",
                        options=destinations,
                        help="Select any available destination airport",
                        key="custom_destination"
                    )
                
                with col3:
                    min_date = schedule_df['Start Date (LZ)'].min()
                    max_date = schedule_df['End Date (LZ)'].max()
                    
                    if pd.notna(min_date) and pd.notna(max_date):
                        custom_date = st.date_input(
                            "Select Shipment Date",
                            value=min_date.date(),
                            min_value=min_date.date(),
                            max_value=max_date.date(),
                            key="custom_date"
                        )
                        
                        day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
                        day_of_week = day_names[custom_date.weekday()]
                
                st.markdown(f"""
                <div style="background-color: #FFF8E8; padding: 15px; border-radius: 5px; border-left: 3px solid #FFB500; margin-top: 20px;">
                    <strong>Selected Custom Route:</strong> {custom_origin} → {custom_destination}<br>
                    <strong>Selected Date:</strong> {custom_date} ({day_of_week})
                </div>
                """, unsafe_allow_html=True)
                
                # Time filter option
                st.markdown("---")
                use_custom_time_filter = st.checkbox(
                    "⏰ Set earliest departure time",
                    help="Only show flights departing after a specific time",
                    key="custom_use_time_filter"
                )
                
                custom_min_departure = 0
                if use_custom_time_filter:
                    st.caption("Show only flights departing from:")
                    col_t1, col_t2 = st.columns(2)
                    with col_t1:
                        custom_dep_hour = st.selectbox(
                            "Hour",
                            options=list(range(24)),
                            format_func=lambda x: f"{x:02d}:00",
                            key="custom_dep_hour"
                        )
                    with col_t2:
                        custom_dep_min = st.selectbox(
                            "Minutes",
                            options=[0, 15, 30, 45],
                            format_func=lambda x: f":{x:02d}",
                            key="custom_dep_min"
                        )
                    custom_min_departure = custom_dep_hour * 60 + custom_dep_min
                    st.info(f"🕐 Showing flights departing **{custom_dep_hour:02d}:{custom_dep_min:02d}** or later")
                
                st.markdown("<br>", unsafe_allow_html=True)
                
                if st.button("🔍 Find Available Routes", type="primary", use_container_width=True, key="custom_search"):
                    if custom_origin and custom_destination:
                        if custom_origin == custom_destination:
                            st.warning("⚠️ Please select different airports for origin and destination.")
                        else:
                            display_route_results(custom_origin, custom_destination, custom_date, schedule_df, min_departure_time=custom_min_departure)
            
            # Tab 3: RadioPharma
            elif selected_tab == "RadioPharma":
                st.markdown("""
                <div style="margin-bottom: 20px;">
                    <div style="font-size: 15px; font-weight: 600; color: #172b4d; margin-bottom: 4px;">RadioPharma Route Finder</div>
                    <div style="font-size: 13px; color: #6b778c;">Specialized routing for radiopharmaceutical shipments with compliance constraints</div>
                </div>
                """, unsafe_allow_html=True)
                
                # Check if RadioPharma config is available
                if rp_config is None:
                    st.warning("""
                    **RadioPharma configuration not found.**
                    
                    The Excel file does not contain the 'RP Info' sheet with RadioPharma routing constraints.
                    Please ensure your Excel file includes the 'RP Info' sheet with:
                    - Prohibited flights (Column A)
                    - Approved Origins (Column D)
                    - Approved Destinations (Column E)
                    """)
                else:
                    # Display dynamic info from config
                    origins_list = ', '.join(sorted(rp_config['approved_origins']))
                    
                    col_rp1, col_rp2, col_rp3, col_rp4 = st.columns(4)
                    with col_rp1:
                        st.markdown(f"""
                        <div class="metric-card" style="text-align: center;">
                            <h3>{len(rp_config['approved_origins'])}</h3>
                            <p>Approved Origins</p>
                        </div>
                        """, unsafe_allow_html=True)
                    with col_rp2:
                        st.markdown(f"""
                        <div class="metric-card" style="text-align: center;">
                            <h3>{len(rp_config['approved_destinations'])}</h3>
                            <p>Approved Destinations</p>
                        </div>
                        """, unsafe_allow_html=True)
                    with col_rp3:
                        st.markdown(f"""
                        <div class="metric-card" style="text-align: center;">
                            <h3>{len(rp_config['prohibited_flights'])}</h3>
                            <p>Prohibited Flights</p>
                        </div>
                        """, unsafe_allow_html=True)
                    with col_rp4:
                        st.markdown(f"""
                        <div class="metric-card" style="text-align: center;">
                            <h3>✓</h3>
                            <p>Compliance Active</p>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    with st.expander("📋 View Approved Origins", expanded=False):
                        st.write(origins_list)
                    
                    st.markdown("<br>", unsafe_allow_html=True)
                    
                    col1, col2, col3 = st.columns(3)
                    
                    with col1:
                        # Filter origins to only approved RadioPharma origins
                        available_rp_origins = sorted([o for o in schedule_df['Orig'].dropna().unique() 
                                                       if o in rp_config['approved_origins']])
                        rp_origin = st.selectbox(
                            "Select Origin Airport",
                            options=available_rp_origins,
                            help="Select from approved RadioPharma origin airports",
                            key="rp_origin"
                        )
                    
                    with col2:
                        # Filter destinations to only approved RadioPharma destinations
                        available_rp_destinations = sorted([d for d in schedule_df['Dest'].dropna().unique() 
                                                            if d in rp_config['approved_destinations']])
                        rp_destination = st.selectbox(
                            "Select Destination Airport",
                            options=available_rp_destinations,
                            help="Select from approved RadioPharma destination airports",
                            key="rp_destination"
                        )
                
                with col3:
                    min_date = schedule_df['Start Date (LZ)'].min()
                    max_date = schedule_df['End Date (LZ)'].max()
                    
                    if pd.notna(min_date) and pd.notna(max_date):
                        rp_date = st.date_input(
                            "Select Shipment Date",
                            value=min_date.date(),
                            min_value=min_date.date(),
                            max_value=max_date.date(),
                            key="rp_date"
                        )
                        
                        day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
                        day_of_week = day_names[rp_date.weekday()]
                
                st.markdown(f"""
                <div style="background-color: #FFF8E8; padding: 15px; border-radius: 5px; border-left: 3px solid #FFB500; margin-top: 20px;">
                    <strong>Selected RadioPharma Route:</strong> {rp_origin} → {rp_destination}<br>
                    <strong>Selected Date:</strong> {rp_date} ({day_of_week})
                </div>
                """, unsafe_allow_html=True)
                
                # Time filter option
                st.markdown("---")
                use_rp_time_filter = st.checkbox(
                    "⏰ Set earliest departure time",
                    help="Only show flights departing after a specific time",
                    key="rp_use_time_filter"
                )
                
                rp_min_departure = 0
                if use_rp_time_filter:
                    st.caption("Show only flights departing from:")
                    col_t1, col_t2 = st.columns(2)
                    with col_t1:
                        rp_dep_hour = st.selectbox(
                            "Hour",
                            options=list(range(24)),
                            format_func=lambda x: f"{x:02d}:00",
                            key="rp_dep_hour"
                        )
                    with col_t2:
                        rp_dep_min = st.selectbox(
                            "Minutes",
                            options=[0, 15, 30, 45],
                            format_func=lambda x: f":{x:02d}",
                            key="rp_dep_min"
                        )
                    rp_min_departure = rp_dep_hour * 60 + rp_dep_min
                    st.info(f"🕐 Showing flights departing **{rp_dep_hour:02d}:{rp_dep_min:02d}** or later")
                
                st.markdown("<br>", unsafe_allow_html=True)
                
                if st.button("🔍 Find RadioPharma Routes", type="primary", use_container_width=True, key="rp_search"):
                    if rp_origin and rp_destination:
                        if rp_origin == rp_destination:
                            st.warning("⚠️ Please select different airports for origin and destination.")
                        else:
                            display_radiopharma_results(rp_origin, rp_destination, rp_date, schedule_df, rp_config, min_departure_time=rp_min_departure)
            
            st.markdown('</div>', unsafe_allow_html=True)  # Close section-card
    
    else:
        # Welcome screen when no file is uploaded
        st.markdown("""
        <div style="text-align: center; padding: 3rem 2rem;">
            <div style="font-size: 4rem; margin-bottom: 1rem;">📦</div>
            <h2 style="color: #351C15; margin-bottom: 0.5rem;">Welcome to UPS Flight Routing System</h2>
            <p style="color: #666; font-size: 1.1rem; margin-bottom: 2rem;">Upload your flight schedule to begin optimizing shipment routes</p>
        </div>
        """, unsafe_allow_html=True)
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.markdown("""
            <div class="metric-card" style="text-align: center; padding: 2rem;">
                <div style="font-size: 2rem; margin-bottom: 0.5rem;">🔗</div>
                <h4 style="margin: 0.5rem 0;">Fewest Stops</h4>
                <p style="font-size: 0.85rem; color: #666;">Minimize handling with fewer connections</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown("""
            <div class="metric-card" style="text-align: center; padding: 2rem;">
                <div style="font-size: 2rem; margin-bottom: 0.5rem;">🚀</div>
                <h4 style="margin: 0.5rem 0;">Fastest Arrival</h4>
                <p style="font-size: 0.85rem; color: #666;">Get packages there as quickly as possible</p>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            st.markdown("""
            <div class="metric-card" style="text-align: center; padding: 2rem;">
                <div style="font-size: 2rem; margin-bottom: 0.5rem;">☢️</div>
                <h4 style="margin: 0.5rem 0;">RadioPharma</h4>
                <p style="font-size: 0.85rem; color: #666;">Compliance-checked routes for sensitive shipments</p>
            </div>
            """, unsafe_allow_html=True)
        
        st.markdown("<br>", unsafe_allow_html=True)
        
        with st.expander("📋 **File Requirements**", expanded=True):
            st.markdown("""
            #### Required Excel Structure
            
            | Sheet Name | Purpose | Required Columns |
            |------------|---------|------------------|
            | **SchedDateLocalTimeFlightSchedul** | Flight schedules | Orig, Dest, Start Date, End Date, DOW, Sched Out/In, Blkhr |
            | **Data** | Route pairs to track | Origin Airport, Destination Airport |
            | **RP Info** *(optional)* | RadioPharma config | Prohibited flights (A), Origins (D), Destinations (E) |
            """)

if __name__ == "__main__":
    main()
